from __future__ import annotations

import hashlib
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import LEGACY_REQUIRED_COLUMNS, MONEY_COLUMNS, REQUIRED_COLUMNS


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file, engine="openpyxl")


def uploaded_file_signature(uploaded_file) -> str:
    return hashlib.sha256(uploaded_file.getvalue()).hexdigest()


def normalize_input_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]

    if all(column in normalized.columns for column in REQUIRED_COLUMNS):
        return normalized

    if all(column in normalized.columns for column in LEGACY_REQUIRED_COLUMNS):
        migrated = pd.DataFrame()
        migrated["Month"] = normalized["Month"]
        migrated["Sales"] = normalized["Revenue"]
        migrated["Direct Costs"] = normalized["COGS"]
        migrated["Labour Cost"] = normalized["Payroll"]
        migrated["Occupancy Cost"] = 0
        migrated["Other Operating Costs"] = (
            pd.to_numeric(normalized["Rent"], errors="coerce")
            + pd.to_numeric(normalized["Marketing"], errors="coerce")
            + pd.to_numeric(normalized["Other Expenses"], errors="coerce")
        )
        return migrated

    return normalized


def build_input_template() -> bytes:
    output = BytesIO()
    columns = REQUIRED_COLUMNS
    instructions = [
        ("Month", "Reporting month. Use one row per month, for example Jan 2026 or 2026-01."),
        ("Sales", "Total business sales for the month, excluding sales tax, GST, or VAT where applicable."),
        ("Direct Costs", "Food ingredients, wholesale merchandise, packaging directly tied to sales, or materials directly associated with sales."),
        ("Labour Cost", "Wages, salaries, superannuation and other directly attributable employee costs where appropriate."),
        ("Occupancy Cost", "Shop rent, market stall fee, site fee, commercial premises cost, licence or recurring location-related cost."),
        ("Other Operating Costs", "Other operating expenses not included above."),
        ("Currency", "Enter all amounts consistently in the same currency. Do not mix currencies within one file."),
        (
            "Opening Cash Balance",
            "The cash available at the beginning of the first reporting month. Enter this separately in the app. Do not add it as a column in the Monthly Data sheet.",
        ),
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template = pd.DataFrame(columns=columns)
        template["Month"] = pd.date_range("2026-01-01", periods=12, freq="MS").strftime("%b %Y")
        template.to_excel(writer, sheet_name="Monthly Data", index=False, startrow=2)
        pd.DataFrame(instructions, columns=["Field", "Definition"]).to_excel(
            writer, sheet_name="Instructions", index=False
        )

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))

        for sheet in workbook.worksheets:
            header_row = 3 if sheet.title == "Monthly Data" else 1
            sheet.freeze_panes = f"A{header_row + 1}"
            for cell in sheet[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                    max(max_length + 3, 14), 55
                )
        monthly = workbook["Monthly Data"]
        monthly["A1"] = "Replace the example months if needed and enter your own financial data."
        monthly["A1"].font = Font(italic=True, color="475569")
        monthly.merge_cells("A1:G1")
        for row in range(4, 16):
            monthly[f"A{row}"].number_format = "yyyy-mm"
            for col in range(2, 8):
                monthly.cell(row=row, column=col).number_format = "$#,##0"
    return output.getvalue()


def validate_and_prepare(df: pd.DataFrame) -> tuple[pd.DataFrame | None, list[str]]:
    errors: list[str] = []
    df = normalize_input_columns(df)

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        errors.append(f"Missing required columns: {', '.join(missing)}.")
        return None, errors

    df = df[REQUIRED_COLUMNS].copy()
    if len(df) < 3:
        errors.append("At least 3 months of financial data are required.")

    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
    if df["Month"].isna().any():
        for index in df[df["Month"].isna()].index:
            errors.append(f"Month contains an unreadable value in row {index + 2}.")

    for column in MONEY_COLUMNS:
        original = df[column]
        missing_cells = original.isna() | (original.astype(str).str.strip() == "")
        for index in df[missing_cells].index:
            errors.append(f"{column} is missing a value in row {index + 2}.")
        df[column] = pd.to_numeric(original, errors="coerce")
        non_numeric = df[column].isna() & ~missing_cells
        for index in df[non_numeric].index:
            errors.append(f"{column} contains a non-numeric value in row {index + 2}.")
        negative = df[column] < 0
        if column == "Sales":
            for index in df[negative].index:
                errors.append(f"Sales cannot be negative in row {index + 2}.")
        else:
            for index in df[negative].index:
                errors.append(f"{column} cannot be negative in row {index + 2}.")

    valid_months = df["Month"].dropna().dt.to_period("M")
    if valid_months.duplicated().any():
        errors.append("Month appears more than once. Please use one row per reporting month.")

    if errors:
        return None, errors

    df = df.sort_values("Month").reset_index(drop=True)
    return df, []
