from __future__ import annotations

from datetime import datetime
from io import BytesIO
from textwrap import wrap

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import business_labels, money, occupancy_percent_label, percent, runway_label


PERCENT_ASSUMPTIONS = {
    "Forecast Sales Growth",
    "Direct Costs %",
    "Labour Cost Growth",
    "Occupancy Cost Growth",
    "Other Operating Costs Growth",
    "Other Operating Expense Growth",
    "Downside Sales Adjustment",
    "Upside Sales Adjustment",
    "Downside Adjustment",
    "Upside Adjustment",
}


def format_excel_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_length + 3, 14), 42
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = str(sheet.cell(row=1, column=cell.column).value or "")
                    row_label = str(sheet.cell(row=cell.row, column=1).value or "")
                    label_or_header = f"{row_label} {header}"
                    if sheet.title == "Assumptions" and row_label in PERCENT_ASSUMPTIONS:
                        cell.number_format = "0.0%"
                    elif any(token in label_or_header for token in ["Margin", "Growth", "%", "Buffer", "Adjustment"]):
                        cell.number_format = "0.0%"
                    elif "Runway" in label_or_header:
                        cell.number_format = "0.0"
                    elif "Score" in label_or_header:
                        cell.number_format = "0"
                    else:
                        cell.number_format = "$#,##0"
                elif sheet.cell(row=1, column=cell.column).value == "Month":
                    cell.number_format = "mmm yyyy"


def prepare_export_dataframe(df: pd.DataFrame, business_type: str) -> pd.DataFrame:
    export_df = df.copy()
    labels = business_labels(business_type)
    rename_map = {
        "Occupancy Cost": labels["occupancy_cost"],
        "Occupancy Cost %": occupancy_percent_label(business_type),
    }
    export_df = export_df.rename(columns=rename_map)
    if business_type != "Food & Beverage":
        export_df = export_df.drop(
            columns=[column for column in ["Prime Cost", "Prime Cost %"] if column in export_df.columns]
        )
    return export_df


def build_scenario_export(
    scenario_details: dict[str, dict[str, pd.DataFrame | dict[str, float]]],
    business_type: str,
) -> pd.DataFrame:
    rows = []
    labels = business_labels(business_type)
    occupancy_label = labels["occupancy_cost"]
    for scenario_name, detail in scenario_details.items():
        scenario_forecast = detail["forecast"]
        scenario_metrics = detail["metrics"]
        rows.append(
            {
                "Scenario": scenario_name,
                "Sales": float(scenario_forecast["Sales"].sum()),
                "Direct Costs": float(scenario_forecast["Direct Costs"].sum()),
                "Labour Cost": float(scenario_forecast["Labour Cost"].sum()),
                occupancy_label: float(scenario_forecast["Occupancy Cost"].sum()),
                "Other Operating Costs": float(scenario_forecast["Other Operating Costs"].sum()),
                "Operating Profit": float(scenario_metrics["operating_profit"]),
                "Ending Cash": float(scenario_metrics["cash_balance"]),
                "Cash Runway": runway_label(
                    float(scenario_metrics["cash_runway"]), float(scenario_metrics["cash_balance"])
                ),
            }
        )
    return pd.DataFrame(rows)


def score_label_from_score(score: int) -> str:
    if score >= 80:
        return "Healthy"
    if score >= 60:
        return "Stable"
    if score >= 40:
        return "Watch"
    return "At Risk"


def build_summary_export(
    history: pd.DataFrame,
    forecast: pd.DataFrame,
    scenarios: pd.DataFrame,
    score_breakdown: pd.DataFrame,
    business_type: str,
    metrics: dict[str, float] | None = None,
    score_label: str | None = None,
) -> pd.DataFrame:
    latest = history.iloc[-1]
    total_sales = float(history["Sales"].sum())
    total_operating_profit = float(history["Operating Profit"].sum())
    historical_operating_margin = total_operating_profit / total_sales if total_sales else 0.0
    total_gross_profit = float(history["Gross Profit"].sum())
    gross_margin = total_gross_profit / total_sales if total_sales else 0.0
    latest_sales = float(latest["Sales"])
    cash_balance = float(latest["Estimated Closing Cash Balance"])
    break_even_sales = float(latest["Break-even Sales"])
    break_even_buffer = (latest_sales - break_even_sales) / latest_sales if latest_sales else 0.0
    score = int(score_breakdown["Score"].sum())
    label = score_label or score_label_from_score(score)
    forecast_ending_cash = float(forecast["Estimated Closing Cash Balance"].iloc[-1])
    base = scenarios.loc[scenarios["Scenario"] == "Base Case"].iloc[0]
    downside = scenarios.loc[scenarios["Scenario"] == "Downside Case"].iloc[0]
    upside = scenarios.loc[scenarios["Scenario"] == "Upside Case"].iloc[0]

    if metrics is not None:
        cash_balance = float(metrics["cash_balance"])
        gross_margin = float(metrics["gross_margin"])
        historical_operating_margin = float(metrics["operating_margin"])
        break_even_sales = float(metrics["break_even_sales"])
        break_even_buffer = float(metrics["break_even_buffer"])

    rows = [
        ("Business Type", business_type),
        ("Financial Health Score", score),
        ("Health Label", label),
        ("Historical Sales", total_sales),
        ("Historical Gross Margin", gross_margin),
        ("Historical Operating Profit", total_operating_profit),
        ("Historical Operating Margin", historical_operating_margin),
        ("Latest Monthly Sales", latest_sales),
        ("Break-even Sales", break_even_sales),
        ("Break-even Buffer", break_even_buffer),
        ("Estimated Cash Balance", cash_balance),
        ("Base Forecast Ending Cash", forecast_ending_cash),
        ("Base Case Sales", float(base["Sales"])),
        ("Downside Case Sales", float(downside["Sales"])),
        ("Upside Case Sales", float(upside["Sales"])),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def export_excel(
    history: pd.DataFrame,
    forecast: pd.DataFrame,
    scenarios: pd.DataFrame,
    score_breakdown: pd.DataFrame,
    assumptions: dict[str, float],
    business_type: str,
    scenario_details: dict[str, dict[str, pd.DataFrame | dict[str, float]]],
    metrics: dict[str, float] | None = None,
    score_label: str | None = None,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        build_summary_export(
            history,
            forecast,
            scenarios,
            score_breakdown,
            business_type,
            metrics,
            score_label,
        ).to_excel(writer, sheet_name="Summary", index=False)
        prepare_export_dataframe(history, business_type).to_excel(
            writer, sheet_name="Historical Analysis", index=False
        )
        prepare_export_dataframe(forecast, business_type).to_excel(
            writer, sheet_name="12-Month Forecast", index=False
        )
        build_scenario_export(scenario_details, business_type).to_excel(
            writer, sheet_name="Scenario Analysis", index=False
        )
        score_breakdown.to_excel(writer, sheet_name="Health Score", index=False)
        pd.DataFrame(
            [{"Assumption": key, "Value": value} for key, value in assumptions.items()]
        ).to_excel(writer, sheet_name="Assumptions", index=False)
        format_excel_workbook(writer)
    return output.getvalue()


def pdf_text_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(
    content: list[str],
    x: float,
    y: float,
    text: str,
    size: int = 10,
    color: tuple[float, float, float] = (0.05, 0.07, 0.10),
) -> None:
    content.append(f"{color[0]:.3f} {color[1]:.3f} {color[2]:.3f} rg")
    content.append("BT")
    content.append(f"/F1 {size} Tf")
    content.append(f"1 0 0 1 {x:.1f} {y:.1f} Tm")
    content.append(f"({pdf_text_escape(text)}) Tj")
    content.append("ET")


def pdf_wrapped_text(
    content: list[str],
    x: float,
    y: float,
    text: str,
    size: int = 10,
    width: int = 86,
    line_height: float = 14,
    color: tuple[float, float, float] = (0.05, 0.07, 0.10),
) -> float:
    for part in wrap(text, width=width) or [""]:
        pdf_text(content, x, y, part, size, color=color)
        y -= line_height
    return y


def pdf_rect(
    content: list[str],
    x: float,
    y: float,
    width: float,
    height: float,
    fill: tuple[float, float, float] | None = None,
    stroke: tuple[float, float, float] | None = (0.82, 0.86, 0.91),
) -> None:
    content.append("q")
    if fill:
        content.append(f"{fill[0]:.3f} {fill[1]:.3f} {fill[2]:.3f} rg")
        content.append(f"{x:.1f} {y:.1f} {width:.1f} {height:.1f} re f")
    if stroke:
        content.append(f"{stroke[0]:.3f} {stroke[1]:.3f} {stroke[2]:.3f} RG")
        content.append(f"{x:.1f} {y:.1f} {width:.1f} {height:.1f} re S")
    content.append("Q")


def pdf_line(
    content: list[str],
    points: list[tuple[float, float]],
    color: tuple[float, float, float] = (0.15, 0.39, 0.92),
    width: float = 1.8,
) -> None:
    if len(points) < 2:
        return
    content.append("q")
    content.append(f"{color[0]:.3f} {color[1]:.3f} {color[2]:.3f} RG")
    content.append(f"{width:.1f} w")
    start_x, start_y = points[0]
    content.append(f"{start_x:.1f} {start_y:.1f} m")
    for x, y in points[1:]:
        content.append(f"{x:.1f} {y:.1f} l")
    content.append("S")
    content.append("Q")


def pdf_cash_balance_chart(
    content: list[str],
    forecast: pd.DataFrame | None,
    x: float,
    y: float,
    width: float,
    height: float,
) -> None:
    pdf_rect(content, x, y, width, height, fill=(1, 1, 1), stroke=(0.86, 0.89, 0.94))
    if forecast is None or "Estimated Closing Cash Balance" not in forecast.columns or forecast.empty:
        pdf_wrapped_text(
            content,
            x + 16,
            y + height - 28,
            "Monthly estimated cash balance trend is available in the Excel export.",
            9,
            width=76,
            color=(0.28, 0.33, 0.40),
        )
        return

    values = [float(value) for value in forecast["Estimated Closing Cash Balance"].head(12)]
    if not values:
        return
    min_value = min(values)
    max_value = max(values)
    if min_value == max_value:
        min_value -= 1
        max_value += 1
    left = x + 44
    bottom = y + 28
    chart_w = width - 72
    chart_h = height - 58
    zero_y = None
    if min_value < 0 < max_value:
        zero_y = bottom + (0 - min_value) / (max_value - min_value) * chart_h
    elif min_value >= 0:
        zero_y = bottom
    else:
        zero_y = bottom + chart_h
    content.append("0.88 0.91 0.95 RG")
    content.append(f"0.7 w {left:.1f} {bottom:.1f} m {left + chart_w:.1f} {bottom:.1f} l S")
    content.append(f"0.7 w {left:.1f} {bottom + chart_h:.1f} m {left + chart_w:.1f} {bottom + chart_h:.1f} l S")
    content.append(f"0.7 w {left:.1f} {zero_y:.1f} m {left + chart_w:.1f} {zero_y:.1f} l S")
    points = []
    denominator = max(len(values) - 1, 1)
    for idx, value in enumerate(values):
        px = left + idx / denominator * chart_w
        py = bottom + (value - min_value) / (max_value - min_value) * chart_h
        points.append((px, py))
    pdf_line(content, points)
    for px, py in points:
        pdf_rect(content, px - 1.8, py - 1.8, 3.6, 3.6, fill=(0.15, 0.39, 0.92), stroke=None)
    pdf_text(content, x + 12, bottom + chart_h - 4, money(max_value), 8, color=(0.28, 0.33, 0.40))
    pdf_text(content, x + 12, bottom - 2, money(min_value), 8, color=(0.28, 0.33, 0.40))
    pdf_text(content, left, y + 10, "Month 1", 8, color=(0.28, 0.33, 0.40))
    pdf_text(content, left + chart_w - 36, y + 10, "Month 12", 8, color=(0.28, 0.33, 0.40))


def pdf_footer(content: list[str], page_number: int, total_pages: int) -> None:
    content.append("0.82 0.86 0.91 RG")
    content.append("54 44 504 0 m S")
    pdf_text(
        content,
        72,
        26,
        f"Business Financial Health Check | Page {page_number} of {total_pages}",
        9,
        color=(0.28, 0.33, 0.40),
    )


def pdf_page_stream(content: list[str], page_number: int, total_pages: int) -> bytes:
    pdf_footer(content, page_number, total_pages)
    return "\n".join(content).encode("latin-1", errors="replace")


def build_report_pdf(
    metrics: dict[str, float],
    score: int,
    score_label: str,
    forecast_metrics: dict[str, float],
    scenarios: pd.DataFrame,
    summary_lines: list[str],
    business_type: str,
    management_priorities: list[str],
    assumptions: dict[str, float | str],
    forecast: pd.DataFrame | None = None,
    score_breakdown: pd.DataFrame | None = None,
) -> bytes:
    generated_date = datetime.now().strftime("%Y-%m-%d")
    pages: list[list[str]] = []

    page1: list[str] = []
    pdf_text(page1, 54, 760, "SENALO", 16, color=(0.12, 0.23, 0.47))
    pdf_text(page1, 54, 736, "Business Financial Health Report", 20)
    pdf_text(page1, 54, 712, "Financial Health and 12-Month Scenario Analysis", 12, color=(0.28, 0.33, 0.40))
    pdf_text(page1, 54, 692, f"Report generated: {generated_date}", 9, color=(0.35, 0.39, 0.46))
    pdf_rect(page1, 54, 592, 504, 78, fill=(0.94, 0.97, 1.0))
    pdf_text(page1, 76, 640, "Financial Health Score", 11, color=(0.23, 0.28, 0.34))
    pdf_text(page1, 76, 610, f"{score}/100", 24)
    pdf_text(page1, 330, 640, "Status", 11, color=(0.23, 0.28, 0.34))
    pdf_text(page1, 330, 610, score_label, 24)
    snapshot = [
        ("Business Type", business_type),
        ("Estimated Cash Balance", money(metrics["cash_balance"])),
        ("Cash Runway", runway_label(metrics["cash_runway"], metrics["cash_balance"])),
        ("Sales", money(metrics["sales"])),
        ("Gross Margin", percent(metrics["gross_margin"])),
        ("Labour Cost %", percent(metrics["labour_cost_percentage"])),
        (occupancy_percent_label(business_type), percent(metrics["occupancy_cost_percentage"])),
        ("Operating Margin", percent(metrics["operating_margin"])),
        ("Break-even Sales", money(metrics["break_even_sales"])),
        ("Break-even Buffer", percent(metrics["break_even_buffer"])),
    ]
    if business_type == "Food & Beverage":
        snapshot.extend(
            [
                ("Food Cost %", percent(metrics["direct_costs_percentage"])),
                ("Prime Cost", money(metrics["prime_cost"])),
                ("Prime Cost %", percent(metrics["prime_cost_percentage"])),
            ]
        )
    x_positions = [54, 226, 398]
    y = 526
    for idx, (label, value) in enumerate(snapshot):
        x = x_positions[idx % 3]
        if idx % 3 == 0 and idx > 0:
            y -= 64
        pdf_rect(page1, x, y, 150, 48, fill=(1, 1, 1))
        pdf_text(page1, x + 14, y + 34, label, 9, color=(0.28, 0.33, 0.40))
        pdf_text(page1, x + 14, y + 13, value, 14)
    y_note = 124
    pdf_text(page1, 54, y_note + 34, "Methodology note", 10, color=(0.23, 0.28, 0.34))
    pdf_wrapped_text(
        page1,
        54,
        y_note + 16,
        "Estimated cash movement uses operating profit as a simplified proxy for operating cash flow. It does not include working capital movements, capital expenditure, financing, tax payments, or owner distributions.",
        8,
        width=112,
        line_height=11,
        color=(0.28, 0.33, 0.40),
    )
    pages.append(page1)

    page2: list[str] = []
    pdf_text(page2, 54, 760, "SENALO", 12, color=(0.12, 0.23, 0.47))
    pdf_text(page2, 54, 738, "12-Month Forecast and Scenario Analysis", 17)
    pdf_rect(page2, 54, 666, 504, 34, fill=(0.12, 0.23, 0.47), stroke=None)
    headers = ["Scenario", "Sales", "Operating Profit", "Estimated Ending Cash", "Cash Runway"]
    cols = [66, 162, 258, 370, 474]
    for x, header in zip(cols, headers):
        pdf_text(page2, x, 678, header, 8, color=(1.0, 1.0, 1.0))
    row_y = 642
    for _, row in scenarios.iterrows():
        pdf_rect(page2, 54, row_y - 8, 504, 22, fill=(1, 1, 1), stroke=(0.90, 0.92, 0.95))
        pdf_text(page2, cols[0], row_y, str(row["Scenario"]), 9)
        pdf_text(page2, cols[1], row_y, money(float(row["Sales"])), 9)
        pdf_text(page2, cols[2], row_y, money(float(row["Operating Profit"])), 9)
        pdf_text(page2, cols[3], row_y, money(float(row["Estimated Ending Cash Balance"])), 9)
        pdf_text(page2, cols[4], row_y, str(row["Cash Runway"]), 9)
        row_y -= 28
    pdf_text(page2, 54, 520, "Estimated Ending Cash by Scenario", 13)
    chart_x, chart_y, chart_w, bar_h = 190, 434, 300, 22
    max_cash = max(abs(float(value)) for value in scenarios["Estimated Ending Cash Balance"]) or 1
    for idx, (_, row) in enumerate(scenarios.iterrows()):
        y_bar = chart_y - idx * 48
        cash_value = float(row["Estimated Ending Cash Balance"])
        width = abs(cash_value) / max_cash * chart_w
        fill = (0.15, 0.39, 0.92) if cash_value >= 0 else (0.86, 0.15, 0.15)
        pdf_text(page2, 66, y_bar + 6, str(row["Scenario"]), 9)
        pdf_rect(page2, chart_x, y_bar, chart_w, bar_h, fill=(0.96, 0.97, 0.98), stroke=None)
        pdf_rect(page2, chart_x, y_bar, width, bar_h, fill=fill, stroke=None)
        pdf_text(page2, chart_x + chart_w + 12, y_bar + 6, money(cash_value), 9)
    pdf_text(page2, 54, 250, "12-Month Estimated Cash Balance Trend", 13)
    pdf_cash_balance_chart(page2, forecast, 54, 82, 504, 150)
    pdf_text(page2, 72, 62, f"Base forecast ending cash: {money(forecast_metrics['cash_balance'])}", 9)
    pages.append(page2)

    risk_lines = [line for line in summary_lines if line.startswith("Primary risk:")]
    action_lines = [line for line in summary_lines if line.startswith("Recommended action:")]
    key_lines = [line for line in summary_lines if line not in risk_lines + action_lines]
    page3: list[str] = []
    pdf_text(page3, 54, 760, "SENALO", 12, color=(0.12, 0.23, 0.47))
    pdf_text(page3, 54, 738, "Financial Summary", 17)
    pdf_text(page3, 54, 704, "Financial Health Score Breakdown", 13)
    pdf_rect(page3, 54, 666, 504, 26, fill=(0.12, 0.23, 0.47), stroke=None)
    breakdown_cols = [66, 232, 286, 350]
    for x, header in zip(breakdown_cols, ["Factor", "Score", "Max", "Reason"]):
        pdf_text(page3, x, 676, header, 8, color=(1, 1, 1))
    y = 642
    if score_breakdown is not None:
        for _, row in score_breakdown.iterrows():
            pdf_rect(page3, 54, y - 8, 504, 28, fill=(1, 1, 1), stroke=(0.90, 0.92, 0.95))
            pdf_text(page3, breakdown_cols[0], y + 4, str(row.get("Factor", ""))[:28], 8)
            pdf_text(page3, breakdown_cols[1], y + 4, str(int(row.get("Score", 0))), 8)
            pdf_text(page3, breakdown_cols[2], y + 4, str(int(row.get("Maximum Score", 0))), 8)
            pdf_wrapped_text(page3, breakdown_cols[3], y + 4, str(row.get("Reason", "")), 7, width=42, line_height=8)
            y -= 34
    pdf_text(page3, 54, y - 6, f"Total score: {score}/100", 10)
    pdf_text(page3, 54, y - 40, "Key Findings", 13)
    y -= 64
    for line in key_lines[:4]:
        y = pdf_wrapped_text(page3, 70, y, f"- {line}", 10, width=92)
        y -= 6
    pdf_text(page3, 54, 268, "Primary Financial Risk", 13)
    pdf_rect(page3, 54, 196, 504, 54, fill=(1.0, 0.95, 0.95), stroke=(0.95, 0.65, 0.65))
    risk_text = risk_lines[0].replace("Primary risk: ", "") if risk_lines else "No primary risk identified."
    pdf_wrapped_text(page3, 76, 232, risk_text, 10, width=92)
    pdf_text(page3, 54, 170, "Management Priorities", 13)
    y_actions = 148
    for index, priority in enumerate(management_priorities[:3], start=1):
        y_actions = pdf_wrapped_text(page3, 70, y_actions, f"{index}. {priority}", 10, width=92)
        y_actions -= 4
    pages.append(page3)

    page4: list[str] = []
    pdf_text(page4, 54, 760, "SENALO", 12, color=(0.12, 0.23, 0.47))
    pdf_text(page4, 54, 738, "Report Notes", 17)
    pdf_text(page4, 54, 700, "Assumptions", 13, color=(0.23, 0.28, 0.34))
    pdf_text(
        page4,
        70,
        674,
        f"Business Type: {assumptions.get('Business Type', business_type)}",
        10,
        color=(0.28, 0.33, 0.40),
    )
    pdf_text(
        page4,
        70,
        654,
        f"Base sales growth: {percent(float(assumptions.get('Forecast Sales Growth', 0.0)))}; Direct cost %: {percent(float(assumptions.get('Direct Costs %', 0.0)))}",
        10,
        color=(0.28, 0.33, 0.40),
    )
    pdf_text(page4, 54, 606, "Limitations and Disclaimer", 13, color=(0.23, 0.28, 0.34))
    pdf_wrapped_text(
        page4,
        54,
        580,
        "This tool provides an indicative financial analysis based on information supplied by the user. It is not accounting, tax, audit, investment, or financial advice. Results should be reviewed alongside actual cash-flow records and professional advice where appropriate.",
        10,
        width=96,
        line_height=13,
        color=(0.28, 0.33, 0.40),
    )
    pages.append(page4)

    streams = [pdf_page_stream(page, index, len(pages)) for index, page in enumerate(pages, start=1)]
    page_objects = []
    content_object_start = 3 + len(streams) + 1
    for idx, _ in enumerate(streams, start=1):
        content_ref = content_object_start + idx - 1
        page_objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 3 0 R >> >> /Contents {content_ref} 0 R >>".encode()
        )

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{' '.join(f'{i} 0 R' for i in range(4, 4 + len(streams)))}] /Count {len(streams)} >>".encode(),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        *page_objects,
        *[
            b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
            for stream in streams
        ],
    ]
    pdf = BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = []
    for index, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f"{index} 0 obj\n".encode())
        pdf.write(obj)
        pdf.write(b"\nendobj\n")
    xref = pdf.tell()
    pdf.write(f"xref\n0 {len(objects) + 1}\n".encode())
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets:
        pdf.write(f"{offset:010d} 00000 n \n".encode())
    pdf.write(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode()
    )
    return pdf.getvalue()
