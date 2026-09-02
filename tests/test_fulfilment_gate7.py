from __future__ import annotations

import io
import logging
import unittest
from dataclasses import replace

from openpyxl import load_workbook

from fulfilment import app as app_module
from fulfilment.analysis_processor import process_order_analysis
from fulfilment.config import Settings
from fulfilment.email_service import RecordingEmailProvider
from fulfilment.fulfilment_service import OrderFulfilmentService
from fulfilment.operator_review import (
    OperatorAuthError,
    ReviewError,
    OperatorIdentity,
    authenticate_operator,
    list_expert_review_orders,
    release_expert_review,
    save_review_draft,
)
from fulfilment.orders import InMemoryOrderStore, process_stripe_event
from fulfilment.result_delivery import (
    ResultDeliveryService,
    is_expert_review_blocked,
    is_expert_review_releasable,
    reissue_result_token,
    reproduce_result_token,
)
from fulfilment.storage import InMemoryUploadStorage
from fulfilment.upload_intake import submit_upload
from tests.test_fulfilment_gate2 import event, session
from tests.test_fulfilment_gate4 import csv_bytes, settings as base_settings, valid_dataframe


def settings() -> Settings:
    source = base_settings()
    return Settings(
        stripe_webhook_secret=source.stripe_webhook_secret,
        stripe_secret_key=source.stripe_secret_key,
        google_cloud_project=source.google_cloud_project,
        full_analysis_price_id=source.full_analysis_price_id,
        expert_review_price_id=source.expert_review_price_id,
        full_analysis_product_id=source.full_analysis_product_id,
        expert_review_product_id=source.expert_review_product_id,
        resend_api_key=source.resend_api_key,
        senalo_email_from=source.senalo_email_from,
        senalo_email_reply_to=source.senalo_email_reply_to,
        senalo_public_fulfilment_base_url=source.senalo_public_fulfilment_base_url,
        token_expiry_days=source.token_expiry_days,
        result_token_expiry_days=source.result_token_expiry_days,
        token_derivation_secret=source.token_derivation_secret,
        upload_bucket=source.upload_bucket,
        max_upload_bytes=source.max_upload_bytes,
        operator_auth_token="operator-secret",
    )


class DummyRequest:
    def __init__(self, token: str | None = None, headers: dict[str, str] | None = None):
        self.query_params = {"t": token} if token is not None else {}
        self.headers = headers or {}


def operator_request() -> DummyRequest:
    return DummyRequest(headers={"x-senalo-operator-token": "operator-secret", "x-senalo-operator-id": "qa-operator"})


def create_validated_order(*, product_code: str = "EXPERT_REVIEW"):
    store = InMemoryOrderStore()
    storage = InMemoryUploadStorage()
    provider = RecordingEmailProvider()
    service = OrderFulfilmentService(store, settings(), provider)
    if product_code == "EXPERT_REVIEW":
        checkout_session = session(
            session_id="cs_gate7_expert",
            price_id="price_expert_test",
            product_id="prod_expert_test",
            amount_total=14900,
            metadata_code="EXPERT_REVIEW",
        )
    else:
        checkout_session = session(
            session_id="cs_gate7_full",
            price_id="price_full_test",
            product_id="prod_full_test",
            amount_total=3900,
            metadata_code="FULL_ANALYSIS",
        )
    process_stripe_event(
        event(f"evt_gate7_{product_code}", "checkout.session.completed", checkout_session),
        store,
        settings(),
        fulfilment_service=service,
    )
    order = next(iter(store.orders.values()))
    uploaded = submit_upload(
        order=order,
        store=store,
        storage=storage,
        business_type="Food & Beverage",
        opening_cash_value="100000",
        filename="financials.csv",
        content=csv_bytes(valid_dataframe()),
        max_upload_bytes=settings().max_upload_bytes,
    )
    return store, storage, uploaded.order


def ready_expert_order():
    store, storage, order = create_validated_order(product_code="EXPERT_REVIEW")
    processed = process_order_analysis(order_id=order.order_id, store=store, storage=storage)
    return store, storage, processed.order


class Gate7ExpertReviewTests(unittest.TestCase):
    def test_operator_auth_is_required_and_customer_token_does_not_authorize(self) -> None:
        with self.assertRaises(OperatorAuthError):
            authenticate_operator(DummyRequest(headers={}), settings())
        with self.assertRaises(OperatorAuthError):
            authenticate_operator(DummyRequest("customer-token", headers={"x-senalo-operator-token": "wrong"}), settings())
        operator = authenticate_operator(operator_request(), settings())
        self.assertEqual(operator.operator_id, "qa-operator")

    def test_only_ready_expert_reviews_appear_in_operator_queue(self) -> None:
        store, storage, expert = ready_expert_order()
        full_store, full_storage, full = create_validated_order(product_code="FULL_ANALYSIS")
        full_ready = process_order_analysis(order_id=full.order_id, store=full_store, storage=full_storage).order
        store.save_order(full_ready)
        reviews = list_expert_review_orders(store)
        self.assertEqual([order.order_id for order in reviews], [expert.order_id])
        self.assertEqual(reviews[0].expert_review_status, "PENDING_REVIEW")

        released = store.save_order(
            replace(
                expert,
                expert_review_status="RELEASED",
                final_pdf_object_path="results/released/final.pdf",
                final_excel_object_path="results/released/final.xlsx",
            )
        )
        self.assertNotIn(released.order_id, [order.order_id for order in list_expert_review_orders(store)])

    def test_draft_saves_human_review_without_releasing_customer_access(self) -> None:
        store, storage, order = ready_expert_order()
        provider = RecordingEmailProvider()
        draft = save_review_draft(
            order_id=order.order_id,
            store=store,
            operator=OperatorIdentity("qa"),
            commentary="<script>review</script>",
            action_values=["Tighten weekly cash reporting", "Review supplier terms", "Prioritise margin-positive sales"],
        )
        self.assertEqual(draft.expert_review_status, "IN_REVIEW")
        self.assertEqual(draft.review_actions[0]["rank"], 1)
        self.assertTrue(is_expert_review_blocked(draft))
        self.assertFalse(is_expert_review_releasable(draft))
        self.assertEqual(ResultDeliveryService(store, settings(), provider).send_result_ready_email(order.order_id).result_email_status, "NOT_SENT")
        self.assertEqual(provider.sent, [])

    def test_release_requires_confirmation_commentary_and_three_to_five_actions(self) -> None:
        store, storage, order = ready_expert_order()
        common = {
            "order_id": order.order_id,
            "store": store,
            "storage": storage,
            "settings": settings(),
            "email_provider": RecordingEmailProvider(),
            "operator": OperatorIdentity("qa"),
            "commentary": "Reviewed against the base analysis.",
            "action_values": ["Action one", "Action two", "Action three"],
        }
        with self.assertRaisesRegex(ReviewError, "REVIEW_CONFIRMATION_REQUIRED"):
            release_expert_review(**common, confirmation=None)
        with self.assertRaisesRegex(ReviewError, "REVIEW_COMMENTARY_REQUIRED"):
            release_expert_review(**{**common, "commentary": " "}, confirmation="approve-release")
        with self.assertRaisesRegex(ReviewError, "REVIEW_ACTIONS_REQUIRED"):
            release_expert_review(**{**common, "action_values": ["Only one"]}, confirmation="approve-release")
        with self.assertRaisesRegex(ReviewError, "REVIEW_ACTIONS_REQUIRED"):
            release_expert_review(**{**common, "action_values": ["Action one", "   ", "Action three"]}, confirmation="approve-release")
        with self.assertRaisesRegex(ReviewError, "REVIEW_ACTIONS_LIMIT"):
            release_expert_review(
                **{**common, "action_values": ["One", "Two", "Three", "Four", "Five", "Six"]},
                confirmation="approve-release",
            )
        with self.assertRaisesRegex(ReviewError, "REVIEW_INPUT_TOO_LONG"):
            release_expert_review(**{**common, "commentary": "x" * 6001}, confirmation="approve-release")
        with self.assertRaisesRegex(ReviewError, "REVIEW_INPUT_TOO_LONG"):
            release_expert_review(**{**common, "action_values": ["One", "Two", "x" * 801]}, confirmation="approve-release")

    def test_three_four_and_five_actions_can_be_approved(self) -> None:
        for count in [3, 4, 5]:
            store, storage, order = ready_expert_order()
            released = release_expert_review(
                order_id=order.order_id,
                store=store,
                storage=storage,
                settings=settings(),
                email_provider=RecordingEmailProvider(),
                operator=OperatorIdentity("qa"),
                commentary=f"Approved with {count} actions.",
                action_values=[f"Action {index}" for index in range(1, count + 1)],
                confirmation="approve-release",
            )
            self.assertEqual(released.expert_review_status, "RELEASED")
            self.assertEqual(len(released.review_actions), count)

    def test_successful_release_creates_final_artifacts_sends_email_and_is_idempotent(self) -> None:
        store, storage, order = ready_expert_order()
        provider = RecordingEmailProvider()
        released = release_expert_review(
            order_id=order.order_id,
            store=store,
            storage=storage,
            settings=settings(),
            email_provider=provider,
            operator=OperatorIdentity("qa"),
            commentary="Sales are stable, but management should protect cash before adding fixed commitments.",
            action_values=["Hold weekly cash review", "Check supplier payment timing", "Protect profitable service mix"],
            confirmation="approve-release",
        )
        self.assertEqual(released.expert_review_status, "RELEASED")
        self.assertIsNotNone(released.released_at)
        self.assertIsNotNone(released.approved_at)
        self.assertEqual(released.review_operator_id, "qa")
        self.assertTrue(released.final_pdf_object_path.startswith(f"results/{order.order_id}/final/"))
        self.assertTrue(released.final_excel_object_path.startswith(f"results/{order.order_id}/final/"))
        final_pdf = storage.load(released.final_pdf_object_path)
        self.assertTrue(final_pdf.startswith(b"%PDF-1.4"))
        self.assertIn(b"Sales are stable", final_pdf)
        self.assertIn(b"1. Hold weekly cash review", final_pdf)
        self.assertIn(b"2. Check supplier payment timing", final_pdf)
        self.assertIn(b"3. Protect profitable service mix", final_pdf)
        self.assertNotIn(b"4.", final_pdf)
        self.assertIn(b"Historical Sales: $33,000", final_pdf)
        self.assertIn(b"Historical Operating Profit: $7,350", final_pdf)
        self.assertIn(b"Historical Operating Margin: 22.3%", final_pdf)
        workbook = load_workbook(io.BytesIO(storage.load(released.final_excel_object_path)), read_only=True)
        self.assertIn("Expert Review", workbook.sheetnames)
        self.assertIn("Historical Analysis", workbook.sheetnames)
        self.assertEqual(len(provider.sent), 1)
        self.assertEqual(provider.sent[0][0].subject, "SENALO Expert Review – Your Review Is Ready")
        self.assertIn("/result?t=", provider.sent[0][0].body)

        second = release_expert_review(
            order_id=order.order_id,
            store=store,
            storage=storage,
            settings=settings(),
            email_provider=provider,
            operator=OperatorIdentity("qa"),
            commentary="Second attempt should not replace approved files.",
            action_values=["One", "Two", "Three"],
            confirmation="approve-release",
        )
        self.assertEqual(second.final_pdf_object_path, released.final_pdf_object_path)
        self.assertEqual(len(provider.sent), 1)

    def test_customer_result_page_is_blocked_before_release_and_final_only_after_release(self) -> None:
        store, storage, order = ready_expert_order()
        _, raw_token = reissue_result_token(
            order.order_id,
            store,
            derivation_secret=settings().token_derivation_secret,
            expiry_days=settings().result_token_expiry_days,
        )
        original_store = app_module.FirestoreOrderStore
        original_settings = app_module.Settings.from_env
        original_storage = app_module.get_upload_storage
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            blocked = app_module.result_page(DummyRequest(raw_token))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertEqual(blocked.status_code, 200)
        self.assertIn("being prepared", blocked.body.decode("utf-8"))
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            pending_pdf = app_module.download_pdf(DummyRequest(raw_token))
            pending_excel = app_module.download_excel(DummyRequest(raw_token))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertEqual(pending_pdf.status_code, 403)
        self.assertEqual(pending_excel.status_code, 403)

        approved = store.save_order(replace(store.get_order(order.order_id), expert_review_status="APPROVED"))
        self.assertTrue(is_expert_review_blocked(approved))
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            approved_pdf = app_module.download_pdf(DummyRequest(raw_token))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertEqual(approved_pdf.status_code, 403)

        provider = RecordingEmailProvider()
        released = release_expert_review(
            order_id=order.order_id,
            store=store,
            storage=storage,
            settings=settings(),
            email_provider=provider,
            operator=OperatorIdentity("qa"),
            commentary="Final review complete.",
            action_values=["Action one", "Action two", "Action three"],
            confirmation="approve-release",
        )
        result_token = reproduce_result_token(released, settings().token_derivation_secret)
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            page = app_module.result_page(DummyRequest(result_token))
            pdf = app_module.download_pdf(DummyRequest(result_token))
            excel = app_module.download_excel(DummyRequest(result_token))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertEqual(page.status_code, 200)
        self.assertIn("SENALO Expert Review", page.body.decode("utf-8"))
        self.assertEqual(pdf.headers["content-disposition"], 'attachment; filename="SENALO-Expert-Review.pdf"')
        self.assertEqual(excel.headers["content-disposition"], 'attachment; filename="SENALO-Expert-Review.xlsx"')
        self.assertNotEqual(pdf.body, storage.load(released.pdf_object_path))
        self.assertNotEqual(excel.body, storage.load(released.excel_object_path))

    def test_optional_replacement_files_are_validated(self) -> None:
        store, storage, order = ready_expert_order()
        base_excel = storage.load(order.excel_object_path)
        provider = RecordingEmailProvider()
        released = release_expert_review(
            order_id=order.order_id,
            store=store,
            storage=storage,
            settings=settings(),
            email_provider=provider,
            operator=OperatorIdentity("qa"),
            commentary="Replacement files reviewed and approved.",
            action_values=["Action one", "Action two", "Action three"],
            confirmation="approve-release",
            replacement_pdf=b"%PDF-1.4\nreplacement\n%%EOF\n",
            replacement_excel=base_excel,
            replacement_excel_filename="review.xlsx",
        )
        self.assertEqual(storage.load(released.final_pdf_object_path), b"%PDF-1.4\nreplacement\n%%EOF\n")
        released_store = store
        released_storage = storage

        store, storage, order = ready_expert_order()
        with self.assertRaisesRegex(ReviewError, "REVIEW_REPLACEMENT_INVALID"):
            release_expert_review(
                order_id=order.order_id,
                store=store,
                storage=storage,
                settings=settings(),
                email_provider=RecordingEmailProvider(),
                operator=OperatorIdentity("qa"),
                commentary="Invalid file.",
                action_values=["Action one", "Action two", "Action three"],
                confirmation="approve-release",
                replacement_pdf=b"not a pdf",
            )

        store, storage, order = ready_expert_order()
        with self.assertRaisesRegex(ReviewError, "REVIEW_REPLACEMENT_INVALID"):
            release_expert_review(
                order_id=order.order_id,
                store=store,
                storage=storage,
                settings=settings(),
                email_provider=RecordingEmailProvider(),
                operator=OperatorIdentity("qa"),
                commentary="Invalid xlsm.",
                action_values=["Action one", "Action two", "Action three"],
                confirmation="approve-release",
                replacement_excel=base_excel,
                replacement_excel_filename="review.xlsm",
            )

        prior_pdf_path = released.final_pdf_object_path
        after_invalid_retry = release_expert_review(
            order_id=released.order_id,
            store=released_store,
            storage=released_storage,
            settings=settings(),
            email_provider=provider,
            operator=OperatorIdentity("qa"),
            commentary="Already released invalid retry should not replace.",
            action_values=["Action one", "Action two", "Action three"],
            confirmation="approve-release",
            replacement_pdf=b"not a pdf",
        )
        self.assertEqual(after_invalid_retry.final_pdf_object_path, prior_pdf_path)

    def test_operator_pages_escape_review_content_and_keep_security_headers(self) -> None:
        store, storage, order = ready_expert_order()
        saved = save_review_draft(
            order_id=order.order_id,
            store=store,
            operator=OperatorIdentity("qa"),
            commentary="<script>alert('x')</script>",
            action_values=["<b>Review payable timing</b>", "Action two", "Action three"],
        )
        original_store = app_module.FirestoreOrderStore
        original_settings = app_module.Settings.from_env
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        try:
            listing = app_module.operator_reviews(operator_request())
            detail = app_module.operator_review_detail(operator_request(), saved.order_id)
            denied = app_module.operator_review_detail(DummyRequest("customer-token"), saved.order_id)
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
        self.assertEqual(listing.status_code, 200)
        self.assertEqual(detail.status_code, 200)
        body = detail.body.decode("utf-8")
        self.assertIn("&lt;script&gt;", body)
        self.assertNotIn("<script>alert", body)
        self.assertIn("&lt;b&gt;Review payable timing&lt;/b&gt;", body)
        self.assertEqual(detail.headers["cache-control"], "no-store")
        self.assertEqual(denied.status_code, 403)

    def test_operator_downloads_require_operator_auth_and_customer_token_cannot_fetch_base_artifacts(self) -> None:
        store, storage, order = ready_expert_order()
        _, result_token = reissue_result_token(
            order.order_id,
            store,
            derivation_secret=settings().token_derivation_secret,
            expiry_days=settings().result_token_expiry_days,
        )
        original_store = app_module.FirestoreOrderStore
        original_settings = app_module.Settings.from_env
        original_storage = app_module.get_upload_storage
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            list_denied = app_module.operator_reviews(DummyRequest())
            detail_denied = app_module.operator_review_detail(DummyRequest(), order.order_id)
            source_denied = app_module.operator_review_download(DummyRequest(), order.order_id, "source")
            base_pdf_denied = app_module.operator_review_download(DummyRequest(), order.order_id, "base-pdf")
            base_excel_denied = app_module.operator_review_download(DummyRequest(), order.order_id, "base-excel")
            customer_source_denied = app_module.operator_review_download(DummyRequest(result_token), order.order_id, "source")
            source_ok = app_module.operator_review_download(operator_request(), order.order_id, "source")
            result_with_operator_credential = app_module.result_page(DummyRequest("operator-secret"))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertEqual(list_denied.status_code, 403)
        self.assertEqual(detail_denied.status_code, 403)
        self.assertEqual(source_denied.status_code, 403)
        self.assertEqual(base_pdf_denied.status_code, 403)
        self.assertEqual(base_excel_denied.status_code, 403)
        self.assertEqual(customer_source_denied.status_code, 403)
        self.assertEqual(source_ok.status_code, 200)
        self.assertEqual(result_with_operator_credential.status_code, 403)

    def test_final_artifact_failure_leaves_customer_blocked_and_logs_do_not_expose_review_text(self) -> None:
        store, storage, order = ready_expert_order()
        failing_storage = InMemoryUploadStorage(fail_save=True)
        failing_storage.objects = dict(storage.objects)
        logs = io.StringIO()
        handler = logging.StreamHandler(logs)
        logger = logging.getLogger("senalo.fulfilment")
        logger.addHandler(handler)
        try:
            with self.assertRaisesRegex(ReviewError, "REVIEW_FINAL_ARTIFACT_FAILED"):
                release_expert_review(
                    order_id=order.order_id,
                    store=store,
                    storage=failing_storage,
                    settings=settings(),
                    email_provider=RecordingEmailProvider(),
                    operator=OperatorIdentity("qa"),
                    commentary="Sensitive human review text 12345",
                    action_values=["Sensitive action one", "Sensitive action two", "Sensitive action three"],
                    confirmation="approve-release",
                )
        finally:
            logger.removeHandler(handler)
        current = store.get_order(order.order_id)
        self.assertEqual(current.expert_review_status, "IN_REVIEW")
        self.assertTrue(is_expert_review_blocked(current))
        self.assertNotIn("Sensitive human review", logs.getvalue())
        self.assertNotIn("Sensitive action", logs.getvalue())

    def test_order_claim_failure_leaves_review_recoverable(self) -> None:
        class ClaimFailureStore(InMemoryOrderStore):
            def claim_expert_review_release(self, order_id, released_at):
                return None

        source_store, storage, order = ready_expert_order()
        store = ClaimFailureStore()
        store.orders = dict(source_store.orders)
        store.checkout_index = dict(source_store.checkout_index)
        with self.assertRaisesRegex(ReviewError, "REVIEW_CONCURRENT_RELEASE"):
            release_expert_review(
                order_id=order.order_id,
                store=store,
                storage=storage,
                settings=settings(),
                email_provider=RecordingEmailProvider(),
                operator=OperatorIdentity("qa"),
                commentary="Claim failure should be recoverable.",
                action_values=["Action one", "Action two", "Action three"],
                confirmation="approve-release",
            )
        self.assertEqual(store.get_order(order.order_id).expert_review_status, "IN_REVIEW")

    def test_full_analysis_delivery_remains_unchanged(self) -> None:
        store, storage, order = create_validated_order(product_code="FULL_ANALYSIS")
        ready = process_order_analysis(order_id=order.order_id, store=store, storage=storage).order
        provider = RecordingEmailProvider()
        sent = ResultDeliveryService(store, settings(), provider).send_result_ready_email(ready.order_id)
        self.assertEqual(sent.result_email_status, "SENT")
        self.assertEqual(provider.sent[0][0].subject, "SENALO Full Analysis – Your Analysis Is Ready")
        token = provider.sent[0][0].body.split("/result?t=", 1)[1].splitlines()[0]
        original_store = app_module.FirestoreOrderStore
        original_settings = app_module.Settings.from_env
        original_storage = app_module.get_upload_storage
        app_module.FirestoreOrderStore = lambda project=None: store
        app_module.Settings.from_env = settings
        app_module.get_upload_storage = lambda active_settings: storage
        try:
            page = app_module.result_page(DummyRequest(token))
            pdf = app_module.download_pdf(DummyRequest(token))
        finally:
            app_module.FirestoreOrderStore = original_store
            app_module.Settings.from_env = original_settings
            app_module.get_upload_storage = original_storage
        self.assertIn("SENALO Full Analysis", page.body.decode("utf-8"))
        self.assertEqual(pdf.headers["content-disposition"], 'attachment; filename="SENALO-Full-Analysis.pdf"')


if __name__ == "__main__":
    unittest.main()
