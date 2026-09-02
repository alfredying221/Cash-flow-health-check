from __future__ import annotations

import io
import logging
import textwrap
from dataclasses import dataclass, replace
from datetime import datetime
from html import escape
from typing import Iterable

from fastapi import Request
from openpyxl import load_workbook

from .config import Settings
from .email_service import EmailProvider
from .models import Order, utc_now
from .orders import OrderStore
from .result_delivery import ResultDeliveryService
from .storage import UploadStorage, UploadStorageError
from .upload_intake import detect_upload_type


logger = logging.getLogger("senalo.fulfilment")

OPERATOR_ACCESS_DENIED = "OPERATOR_ACCESS_DENIED"
REVIEW_NOT_FOUND = "REVIEW_NOT_FOUND"
REVIEW_NOT_READY = "REVIEW_NOT_READY"
REVIEW_CONFIRMATION_REQUIRED = "REVIEW_CONFIRMATION_REQUIRED"
REVIEW_COMMENTARY_REQUIRED = "REVIEW_COMMENTARY_REQUIRED"
REVIEW_ACTIONS_REQUIRED = "REVIEW_ACTIONS_REQUIRED"
REVIEW_ACTIONS_LIMIT = "REVIEW_ACTIONS_LIMIT"
REVIEW_INPUT_TOO_LONG = "REVIEW_INPUT_TOO_LONG"
REVIEW_FINAL_ARTIFACT_FAILED = "REVIEW_FINAL_ARTIFACT_FAILED"
REVIEW_CONCURRENT_RELEASE = "REVIEW_CONCURRENT_RELEASE"
REVIEW_REPLACEMENT_INVALID = "REVIEW_REPLACEMENT_INVALID"

MAX_COMMENTARY_CHARS = 6000
MAX_ACTION_CHARS = 800
MIN_RELEASE_ACTIONS = 3
MAX_RELEASE_ACTIONS = 5


class OperatorAuthError(Exception):
    pass


class ReviewError(Exception):
    def __init__(self, error_code: str) -> None:
        super().__init__(error_code)
        self.error_code = error_code


@dataclass(frozen=True)
class OperatorIdentity:
    operator_id: str


@dataclass(frozen=True)
class FinalArtifactSet:
    pdf: bytes
    excel: bytes


def authenticate_operator(request: Request, settings: Settings) -> OperatorIdentity:
    configured_token = settings.operator_auth_token
    supplied_token = request.headers.get("x-senalo-operator-token")
    if not configured_token or not supplied_token or supplied_token != configured_token:
        raise OperatorAuthError(OPERATOR_ACCESS_DENIED)
    operator_id = request.headers.get("x-senalo-operator-id") or "operator"
    return OperatorIdentity(operator_id=operator_id[:80])


def list_expert_review_orders(store: OrderStore) -> list[Order]:
    reviews = [
        order
        for order in store.list_orders()
        if order.product_code == "EXPERT_REVIEW"
        and order.fulfilment_status == "READY"
        and order.result_status == "READY"
        and order.expert_review_status in {"PENDING_REVIEW", "IN_REVIEW", "APPROVED"}
    ]
    return sorted(reviews, key=lambda order: order.updated_at, reverse=True)


def get_review_order(store: OrderStore, order_id: str) -> Order:
    order = store.get_order(order_id)
    if not order or order.product_code != "EXPERT_REVIEW":
        raise ReviewError(REVIEW_NOT_FOUND)
    return order


def validate_review_ready(order: Order) -> None:
    if (
        order.payment_status != "PAID"
        or order.fulfilment_status != "READY"
        or order.result_status != "READY"
        or not order.pdf_object_path
        or not order.excel_object_path
    ):
        raise ReviewError(REVIEW_NOT_READY)


def parse_actions(values: Iterable[str | None]) -> list[dict[str, str | int]]:
    actions = []
    for value in values:
        raw_text = value or ""
        if raw_text and not raw_text.strip():
            raise ReviewError(REVIEW_ACTIONS_REQUIRED)
        text = raw_text.strip()
        if not text:
            continue
        if len(text) > MAX_ACTION_CHARS:
            raise ReviewError(REVIEW_INPUT_TOO_LONG)
        actions.append({"rank": len(actions) + 1, "text": text})
    if len(actions) > MAX_RELEASE_ACTIONS:
        raise ReviewError(REVIEW_ACTIONS_LIMIT)
    return actions


def validate_commentary(commentary: str) -> str:
    cleaned = commentary.strip()
    if len(cleaned) > MAX_COMMENTARY_CHARS:
        raise ReviewError(REVIEW_INPUT_TOO_LONG)
    return cleaned


def validate_release_inputs(commentary: str, actions: list[dict[str, str | int]]) -> None:
    if not commentary:
        raise ReviewError(REVIEW_COMMENTARY_REQUIRED)
    if len(actions) < MIN_RELEASE_ACTIONS or len(actions) > MAX_RELEASE_ACTIONS:
        raise ReviewError(REVIEW_ACTIONS_REQUIRED)


def save_review_draft(
    *,
    order_id: str,
    store: OrderStore,
    operator: OperatorIdentity,
    commentary: str,
    action_values: Iterable[str | None],
    now: datetime | None = None,
) -> Order:
    order = get_review_order(store, order_id)
    validate_review_ready(order)
    cleaned_commentary = validate_commentary(commentary)
    actions = parse_actions(action_values)
    saved_at = now or utc_now()
    updated = replace(
        order,
        expert_review_status="IN_REVIEW" if order.expert_review_status == "PENDING_REVIEW" else order.expert_review_status,
        review_commentary=cleaned_commentary,
        review_actions=actions,
        review_started_at=order.review_started_at or saved_at,
        review_updated_at=saved_at,
        review_operator_id=operator.operator_id,
        updated_at=saved_at,
    )
    logger.info("expert_review_draft_saved", extra={"order_id": order.order_id, "operator_id": operator.operator_id})
    return store.save_order(updated)


def release_expert_review(
    *,
    order_id: str,
    store: OrderStore,
    storage: UploadStorage,
    settings: Settings,
    email_provider: EmailProvider,
    operator: OperatorIdentity,
    commentary: str,
    action_values: Iterable[str | None],
    confirmation: str | None,
    replacement_pdf: bytes | None = None,
    replacement_excel: bytes | None = None,
    replacement_excel_filename: str | None = None,
    now: datetime | None = None,
) -> Order:
    confirmed = str(confirmation or "").lower() in {"true", "yes", "on", "approve-release", "release"}
    if not confirmed:
        raise ReviewError(REVIEW_CONFIRMATION_REQUIRED)

    order = get_review_order(store, order_id)
    if order.expert_review_status == "RELEASED":
        return order
    validate_review_ready(order)
    cleaned_commentary = validate_commentary(commentary)
    actions = parse_actions(action_values)
    validate_release_inputs(cleaned_commentary, actions)

    release_started_at = now or utc_now()
    in_review = replace(
        order,
        expert_review_status="IN_REVIEW" if order.expert_review_status == "PENDING_REVIEW" else order.expert_review_status,
        review_commentary=cleaned_commentary,
        review_actions=actions,
        review_started_at=order.review_started_at or release_started_at,
        review_updated_at=release_started_at,
        review_operator_id=operator.operator_id,
        updated_at=release_started_at,
    )
    store.save_order(in_review)

    claimed = store.claim_expert_review_release(order_id, release_started_at)
    if not claimed:
        current = store.get_order(order_id)
        if current and current.expert_review_status == "RELEASED":
            return current
        raise ReviewError(REVIEW_CONCURRENT_RELEASE)

    try:
        final_artifacts = build_final_artifacts(
            order=claimed,
            storage=storage,
            commentary=cleaned_commentary,
            actions=actions,
            replacement_pdf=replacement_pdf,
            replacement_excel=replacement_excel,
            replacement_excel_filename=replacement_excel_filename,
        )
        pdf_result = storage.save_final_result(
            order_id=claimed.order_id,
            content=final_artifacts.pdf,
            extension=".pdf",
            content_type="application/pdf",
        )
        excel_result = storage.save_final_result(
            order_id=claimed.order_id,
            content=final_artifacts.excel,
            extension=".xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReviewError:
        rollback = replace(claimed, expert_review_status="IN_REVIEW", approved_at=None, updated_at=utc_now())
        store.save_order(rollback)
        raise
    except Exception as exc:
        rollback = replace(claimed, expert_review_status="IN_REVIEW", approved_at=None, updated_at=utc_now())
        store.save_order(rollback)
        raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED) from exc

    released_at = utc_now()
    released = replace(
        claimed,
        expert_review_status="RELEASED",
        review_commentary=cleaned_commentary,
        review_actions=actions,
        review_updated_at=released_at,
        review_operator_id=operator.operator_id,
        released_at=released_at,
        final_pdf_object_path=pdf_result.object_path,
        final_pdf_size_bytes=pdf_result.size_bytes,
        final_excel_object_path=excel_result.object_path,
        final_excel_size_bytes=excel_result.size_bytes,
        updated_at=released_at,
    )
    saved = store.save_order(released)
    ResultDeliveryService(store, settings, email_provider).send_result_ready_email(saved.order_id)
    logger.info("expert_review_released", extra={"order_id": saved.order_id, "operator_id": operator.operator_id})
    return store.get_order(saved.order_id) or saved


def build_final_artifacts(
    *,
    order: Order,
    storage: UploadStorage,
    commentary: str,
    actions: list[dict[str, str | int]],
    replacement_pdf: bytes | None,
    replacement_excel: bytes | None,
    replacement_excel_filename: str | None,
) -> FinalArtifactSet:
    if replacement_pdf is not None:
        validate_replacement_pdf(replacement_pdf)
        pdf = replacement_pdf
    else:
        try:
            base_excel = storage.load(str(order.excel_object_path))
        except UploadStorageError as exc:
            raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED) from exc
        pdf = build_expert_review_pdf(order, commentary, actions, base_excel)

    if replacement_excel is not None:
        validate_replacement_excel(replacement_excel_filename or "replacement.xlsx", replacement_excel)
        excel = replacement_excel
    else:
        try:
            base_excel = storage.load(str(order.excel_object_path))
        except UploadStorageError as exc:
            raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED) from exc
        excel = build_expert_review_excel(base_excel, commentary, actions)
    return FinalArtifactSet(pdf=pdf, excel=excel)


def validate_replacement_pdf(content: bytes) -> None:
    if not content.startswith(b"%PDF-"):
        raise ReviewError(REVIEW_REPLACEMENT_INVALID)


def validate_replacement_excel(filename: str, content: bytes) -> None:
    try:
        extension, _ = detect_upload_type(filename, content)
    except Exception as exc:
        raise ReviewError(REVIEW_REPLACEMENT_INVALID) from exc
    if extension != ".xlsx":
        raise ReviewError(REVIEW_REPLACEMENT_INVALID)
    try:
        load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise ReviewError(REVIEW_REPLACEMENT_INVALID) from exc


def build_expert_review_excel(base_excel: bytes, commentary: str, actions: list[dict[str, str | int]]) -> bytes:
    workbook = load_workbook(io.BytesIO(base_excel))
    if "Expert Review" in workbook.sheetnames:
        del workbook["Expert Review"]
    sheet = workbook.create_sheet("Expert Review", 0)
    sheet["A1"] = "SENALO Expert Review"
    sheet["A3"] = "Customised Commentary"
    sheet["A4"] = commentary
    sheet["A6"] = "Prioritised Management Actions"
    for index, action in enumerate(actions, start=7):
        sheet[f"A{index}"] = action.get("rank")
        sheet[f"B{index}"] = action.get("text")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_expert_review_pdf(order: Order, commentary: str, actions: list[dict[str, str | int]], base_excel: bytes) -> bytes:
    canonical_lines = canonical_financial_summary_lines(base_excel)
    lines = [
        "SENALO Expert Review",
        "Human-reviewed final report",
        "",
        "This file includes the Full Analysis calculation summary and SENALO's customised human review.",
        "",
        "Canonical Financial Summary",
        *canonical_lines,
        "",
        "Customised Commentary",
        *wrap_text(commentary, 88),
        "",
        "Prioritised Management Actions",
    ]
    for action in actions:
        lines.extend(wrap_text(f"{action.get('rank')}. {action.get('text')}", 88))
    lines.extend(["", f"Order reference: {order.order_id}", "SENALO | See clearly. Decide better."])
    return simple_pdf(lines)


def canonical_financial_summary_lines(base_excel: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(base_excel), data_only=True, read_only=True)
    if "Historical Analysis" not in workbook.sheetnames:
        raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED)
    sheet = workbook["Historical Analysis"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED)
    headers = [str(value) if value is not None else "" for value in rows[0]]
    data_rows = [row for row in rows[1:] if row and row[1] is not None]
    if not data_rows:
        raise ReviewError(REVIEW_FINAL_ARTIFACT_FAILED)

    def column_total(name: str) -> float:
        index = headers.index(name)
        return float(sum((row[index] or 0) for row in data_rows))

    def latest_value(name: str) -> float:
        index = headers.index(name)
        return float(data_rows[-1][index] or 0)

    total_sales = column_total("Sales")
    total_operating_profit = column_total("Operating Profit")
    operating_margin = total_operating_profit / total_sales if total_sales else 0.0
    latest_sales = latest_value("Sales")
    latest_cash = latest_value("Estimated Closing Cash Balance")
    latest_break_even = latest_value("Break-even Sales")
    return [
        f"Historical Sales: {format_currency(total_sales)}",
        f"Historical Operating Profit: {format_currency(total_operating_profit)}",
        f"Historical Operating Margin: {format_percentage(operating_margin)}",
        f"Latest Sales: {format_currency(latest_sales)}",
        f"Latest Estimated Closing Cash Balance: {format_currency(latest_cash)}",
        f"Latest Break-even Sales: {format_currency(latest_break_even)}",
    ]


def format_currency(value: float) -> str:
    return f"${value:,.0f}"


def format_percentage(value: float) -> str:
    return f"{value * 100:.1f}%"


def wrap_text(text: str, width: int) -> list[str]:
    wrapped: list[str] = []
    for paragraph in text.splitlines() or [""]:
        wrapped.extend(textwrap.wrap(paragraph, width=width) or [""])
    return wrapped


def simple_pdf(lines: list[str]) -> bytes:
    escaped_lines = [line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)") for line in lines]
    content_lines = ["BT", "/F1 11 Tf", "50 790 Td", "14 TL"]
    first = True
    for line in escaped_lines:
        if not first:
            content_lines.append("T*")
        content_lines.append(f"({line}) Tj")
        first = False
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_start = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode("ascii")
    )
    return bytes(output)


def safe_review_text(value: object) -> str:
    return escape(str(value or ""))
