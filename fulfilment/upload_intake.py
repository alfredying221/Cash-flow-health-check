from __future__ import annotations

import io
import math
import re
import zipfile
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from senalo_analysis import (
    BUSINESS_TYPES,
    LEGACY_REQUIRED_COLUMNS,
    MONEY_COLUMNS,
    read_uploaded_file,
    validate_and_prepare,
)

from .models import Order, utc_now
from .orders import OrderStore
from .storage import UploadStorage, UploadStorageError


ALLOWED_UPLOAD_STATUSES = {"AWAITING_UPLOAD", "VALIDATION_FAILED", "VALIDATED"}
MAX_FILENAME_LENGTH = 120
FORMULA_ERROR_MESSAGE = "Please replace formulas in the financial data columns with values before uploading."
LEGACY_MONEY_COLUMNS = ["Revenue", "COGS", "Payroll", "Rent", "Marketing", "Other Expenses"]


class IntakeError(Exception):
    error_code = "INTAKE_ERROR"


class UploadNotAllowedError(IntakeError):
    error_code = "UPLOAD_NOT_ALLOWED"


class InvalidBusinessTypeError(IntakeError):
    error_code = "INVALID_BUSINESS_TYPE"


class InvalidOpeningCashError(IntakeError):
    error_code = "INVALID_OPENING_CASH"


class InvalidUploadError(IntakeError):
    error_code = "INVALID_UPLOAD"


class ValidationFailedError(IntakeError):
    error_code = "VALIDATION_FAILED"

    def __init__(self, errors: list[str]) -> None:
        super().__init__("; ".join(errors))
        self.errors = errors


@dataclass(frozen=True)
class IntakeResult:
    order: Order
    validation_errors: list[str]


class NamedBytesIO(io.BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


def sanitize_filename(filename: str) -> str:
    name = Path(filename or "upload").name
    name = re.sub(r"[^A-Za-z0-9._ -]", "_", name).strip()
    if not name:
        return "upload"
    return name[:MAX_FILENAME_LENGTH]


def parse_opening_cash(value: object) -> float:
    try:
        opening_cash = float(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise InvalidOpeningCashError("Opening Cash must be a valid number.") from exc
    if not math.isfinite(opening_cash) or opening_cash < 0:
        raise InvalidOpeningCashError("Opening Cash must be zero or a positive number.")
    return opening_cash


def validate_business_type(value: str) -> str:
    if value not in BUSINESS_TYPES:
        raise InvalidBusinessTypeError("Select a valid Business Type.")
    return value


def detect_upload_type(filename: str, content: bytes) -> tuple[str, str]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xlsm":
        raise InvalidUploadError("Macro-enabled Excel files are not supported.")
    if suffix == ".xls":
        raise InvalidUploadError("Only CSV and XLSX files are supported.")
    if suffix not in {".csv", ".xlsx"}:
        raise InvalidUploadError("Only CSV and XLSX files are supported.")
    if not content:
        raise InvalidUploadError("Uploaded file is empty.")
    if suffix == ".xlsx":
        if not zipfile.is_zipfile(io.BytesIO(content)):
            raise InvalidUploadError("The XLSX file could not be read. Please upload a valid workbook.")
        return ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if content.startswith(b"MZ") or b"\x00" in content[:2048]:
        raise InvalidUploadError("Uploaded file type is not supported.")
    return ".csv", "text/csv"


def parse_and_validate_financial_file(filename: str, content: bytes) -> list[str]:
    extension, _ = detect_upload_type(filename, content)
    reject_formula_cells(filename, content, extension)
    uploaded = NamedBytesIO(content, filename)
    try:
        if extension == ".csv":
            dataframe = read_uploaded_file(uploaded)
        else:
            dataframe = read_uploaded_file(uploaded)
    except Exception:
        raise InvalidUploadError("The uploaded file could not be read. Please check the file format.")

    prepared, errors = validate_and_prepare(dataframe)
    if errors:
        raise ValidationFailedError(errors)
    if prepared is None:
        raise ValidationFailedError(["The uploaded file could not be validated."])
    return []


def reject_formula_cells(filename: str, content: bytes, extension: str | None = None) -> None:
    active_extension = extension or Path(filename or "").suffix.lower()
    if active_extension != ".xlsx":
        return
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise InvalidUploadError("The XLSX file could not be read. Please upload a valid workbook.") from exc

    try:
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1), [])
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
        columns_to_check: list[str]
        if all(column in headers for column in MONEY_COLUMNS):
            columns_to_check = MONEY_COLUMNS
        elif all(column in headers for column in LEGACY_REQUIRED_COLUMNS):
            columns_to_check = LEGACY_MONEY_COLUMNS
        else:
            return

        column_indexes = [headers.index(column) + 1 for column in columns_to_check]
        for row in worksheet.iter_rows(min_row=2):
            for column_index in column_indexes:
                if column_index <= len(row):
                    value = row[column_index - 1].value
                    if isinstance(value, str) and value.strip().startswith("="):
                        raise ValidationFailedError([FORMULA_ERROR_MESSAGE])
    finally:
        workbook.close()


def ensure_upload_allowed(order: Order) -> None:
    if order.payment_status != "PAID" or order.fulfilment_status not in ALLOWED_UPLOAD_STATUSES:
        raise UploadNotAllowedError("This order is not eligible for upload.")


def mark_validation_failed(
    order: Order,
    store: OrderStore,
    *,
    business_type: str | None = None,
    opening_cash: float | None = None,
    error_code: str,
    now: datetime | None = None,
) -> Order:
    current_time = now or utc_now()
    updated = replace(
        order,
        business_type=business_type or order.business_type,
        opening_cash=opening_cash if opening_cash is not None else order.opening_cash,
        fulfilment_status="VALIDATION_FAILED",
        validation_status="FAILED",
        validation_error_code=error_code,
        updated_at=current_time,
    )
    return store.save_order(updated)


def submit_upload(
    *,
    order: Order,
    store: OrderStore,
    storage: UploadStorage,
    business_type: str,
    opening_cash_value: object,
    filename: str,
    content: bytes,
    max_upload_bytes: int,
) -> IntakeResult:
    ensure_upload_allowed(order)
    selected_business_type = validate_business_type(business_type)
    opening_cash = parse_opening_cash(opening_cash_value)

    if len(content) > max_upload_bytes:
        failed = mark_validation_failed(
            order,
            store,
            business_type=selected_business_type,
            opening_cash=opening_cash,
            error_code="UPLOAD_TOO_LARGE",
        )
        raise ValidationFailedError([f"Uploaded file is too large. Maximum size is {max_upload_bytes} bytes."])

    try:
        extension, content_type = detect_upload_type(filename, content)
        parse_and_validate_financial_file(filename, content)
    except ValidationFailedError as exc:
        mark_validation_failed(
            order,
            store,
            business_type=selected_business_type,
            opening_cash=opening_cash,
            error_code="VALIDATION_FAILED",
        )
        raise exc
    except InvalidUploadError as exc:
        mark_validation_failed(
            order,
            store,
            business_type=selected_business_type,
            opening_cash=opening_cash,
            error_code=exc.error_code,
        )
        raise exc

    previous_object = order.upload_object_path
    stored = storage.save(
        order_id=order.order_id,
        content=content,
        extension=extension,
        content_type=content_type,
    )

    current_time = utc_now()
    updated = replace(
        order,
        business_type=selected_business_type,
        opening_cash=opening_cash,
        upload_status="UPLOADED",
        upload_object_path=stored.object_path,
        upload_original_filename=sanitize_filename(filename),
        upload_content_type=content_type,
        upload_size_bytes=stored.size_bytes,
        upload_created_at=current_time,
        validation_status="VALIDATED",
        validation_error_code=None,
        validated_at=current_time,
        fulfilment_status="VALIDATED",
        updated_at=current_time,
    )

    try:
        saved = store.save_order(updated)
    except Exception:
        try:
            storage.delete(stored.object_path)
        finally:
            raise

    if previous_object and previous_object != stored.object_path:
        try:
            storage.delete(previous_object)
        except UploadStorageError:
            pass

    return IntakeResult(order=saved, validation_errors=[])
