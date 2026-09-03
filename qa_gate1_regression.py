from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from senalo_analysis import (
    build_forecast,
    build_report_pdf,
    build_scenario_from_base,
    calculate_financials,
    export_excel,
    make_cfo_summary,
    make_management_priorities,
    read_uploaded_file,
    runway_label,
    sales_growth_rate,
    score_cash_health,
    summarize_metrics,
    validate_and_prepare,
)


ROOT = Path(__file__).resolve().parent


class NamedBytesIO(BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


CASES = [
    {
        "file": "healthy_food_business.csv",
        "business_type": "Food & Beverage",
        "opening_cash": 100000,
        "expected": {
            "sales": 1122000.0,
            "gross_margin": 0.65,
            "labour_pct": 0.25882352941176473,
            "occupancy_pct": 0.09197860962566845,
            "operating_profit": 250800.0,
            "operating_margin": 0.2235294117647059,
            "sales_trend": 0.2804878048780488,
            "break_even_sales": 64230.769230769234,
            "break_even_buffer": 0.3882783882783883,
            "cash_balance": 350800.0,
            "runway": "Cash Generating",
            "score": 100,
            "label": "Healthy",
        },
    },
    {
        "file": "watch_food_business.csv",
        "business_type": "Food & Beverage",
        "opening_cash": 60000,
        "expected": {"score": 42, "label": "Watch"},
    },
    {
        "file": "at_risk_food_business.csv",
        "business_type": "Food & Beverage",
        "opening_cash": 50000,
        "expected": {"score": 0, "label": "At Risk"},
    },
    {
        "file": "healthy_market_vendor.csv",
        "business_type": "Market Stall / Vendor",
        "opening_cash": 100000,
        "expected": {"score": 100, "label": "Healthy"},
    },
    {
        "file": "watch_market_vendor.csv",
        "business_type": "Market Stall / Vendor",
        "opening_cash": 60000,
        "expected": {"score": 45, "label": "Watch"},
    },
    {
        "file": "sample_data.csv",
        "business_type": "Food & Beverage",
        "opening_cash": 50000,
        "expected": {
            "operating_margin": 0.09800829875518673,
            "break_even_buffer": 0.25413223140495866,
            "score": 90,
            "label": "Healthy",
        },
    },
]


def assert_close(name: str, actual: float, expected: float, tolerance: float = 1e-9) -> None:
    if abs(actual - expected) > tolerance:
        raise AssertionError(f"{name}: expected {expected}, got {actual}")


def load_case(file_name: str, opening_cash: float):
    df = pd.read_csv(ROOT / file_name)
    prepared, errors = validate_and_prepare(df)
    if errors:
        raise AssertionError(f"{file_name} validation failed: {errors}")
    history = calculate_financials(prepared, opening_cash)
    metrics = summarize_metrics(history)
    score, label, breakdown = score_cash_health(history, metrics)
    return history, metrics, score, label, breakdown


def scenario_rows(history: pd.DataFrame, metrics: dict[str, float]):
    forecast = build_forecast(
        history,
        metrics["cash_balance"],
        0.0,
        metrics["direct_costs_percentage"],
        0.0,
        0.0,
        0.0,
    )
    scenario_details = {}
    rows = []
    for name, adjustment in [
        ("Base Case", 0.0),
        ("Downside Case", -0.15),
        ("Upside Case", 0.15),
    ]:
        scenario, row = build_scenario_from_base(
            forecast,
            metrics["cash_balance"],
            name,
            adjustment,
            metrics["direct_costs_percentage"],
        )
        scenario_details[name] = {
            "forecast": scenario,
            "metrics": summarize_metrics(scenario),
        }
        rows.append(row)
    return forecast, pd.DataFrame(rows), scenario_details


def run() -> None:
    results = []
    for case in CASES:
        history, metrics, score, label, breakdown = load_case(case["file"], case["opening_cash"])
        expected = case["expected"]

        checks = {
            "sales": metrics["sales"],
            "gross_margin": metrics["gross_margin"],
            "labour_pct": metrics["labour_cost_percentage"],
            "occupancy_pct": metrics["occupancy_cost_percentage"],
            "operating_profit": metrics["operating_profit"],
            "operating_margin": metrics["operating_margin"],
            "sales_trend": sales_growth_rate(history),
            "break_even_sales": metrics["break_even_sales"],
            "break_even_buffer": metrics["break_even_buffer"],
            "cash_balance": metrics["cash_balance"],
        }
        for key, actual in checks.items():
            if key in expected:
                assert_close(f"{case['file']} {key}", actual, expected[key])
        if "runway" in expected and runway_label(metrics["cash_runway"], metrics["cash_balance"]) != expected["runway"]:
            raise AssertionError(f"{case['file']} runway label changed")
        if score != expected["score"]:
            raise AssertionError(f"{case['file']} score expected {expected['score']}, got {score}")
        if label != expected["label"]:
            raise AssertionError(f"{case['file']} label expected {expected['label']}, got {label}")
        if int(breakdown["Score"].sum()) != score:
            raise AssertionError(f"{case['file']} score components do not equal score")

        forecast, scenarios, scenario_details = scenario_rows(history, metrics)
        base_sales = float(scenarios.loc[scenarios["Scenario"] == "Base Case", "Sales"].iloc[0])
        downside_sales = float(scenarios.loc[scenarios["Scenario"] == "Downside Case", "Sales"].iloc[0])
        upside_sales = float(scenarios.loc[scenarios["Scenario"] == "Upside Case", "Sales"].iloc[0])
        assert_close(f"{case['file']} downside/base", downside_sales / base_sales, 0.85)
        assert_close(f"{case['file']} upside/base", upside_sales / base_sales, 1.15)

        summary = make_cfo_summary(
            history,
            metrics,
            label,
            scenario_details["Downside Case"]["metrics"],
        )
        priorities = make_management_priorities(
            history,
            metrics,
            scenario_details["Downside Case"]["metrics"],
        )
        if len(priorities) != 3:
            raise AssertionError(f"{case['file']} management priorities count changed")

        assumptions = {
            "Business Type": case["business_type"],
            "Forecast Sales Growth": 0.0,
            "Direct Costs %": metrics["direct_costs_percentage"],
            "Labour Growth": 0.0,
            "Occupancy Growth": 0.0,
            "Other Operating Cost Growth": 0.0,
        }
        excel_bytes = export_excel(
            history,
            forecast,
            scenarios,
            breakdown,
            assumptions,
            case["business_type"],
            scenario_details,
            metrics,
            label,
        )
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)
        expected_sheets = [
            "Summary",
            "Historical Analysis",
            "12-Month Forecast",
            "Scenario Analysis",
            "Health Score",
            "Assumptions",
        ]
        if workbook.sheetnames != expected_sheets:
            raise AssertionError(f"{case['file']} Excel sheets changed: {workbook.sheetnames}")

        pdf_bytes = build_report_pdf(
            metrics,
            score,
            label,
            summarize_metrics(forecast),
            scenarios,
            summary,
            case["business_type"],
            priorities,
            assumptions,
            forecast,
            breakdown,
        )
        for text in [b"SENALO", b"Business Financial Health Report", b"Financial Summary", b"Management Priorities"]:
            if text not in pdf_bytes:
                raise AssertionError(f"{case['file']} PDF missing {text.decode()}")

        results.append(
            {
                "file": case["file"],
                "score": score,
                "label": label,
                "operating_margin": metrics["operating_margin"],
                "break_even_buffer": metrics["break_even_buffer"],
                "component_sum": int(breakdown["Score"].sum()),
                "downside_base": downside_sales / base_sales,
                "upside_base": upside_sales / base_sales,
                "pdf_bytes": len(pdf_bytes),
                "excel_bytes": len(excel_bytes),
            }
        )

    legacy = pd.read_csv(ROOT / "sample_data.csv").rename(
        columns={
            "Sales": "Revenue",
            "Direct Costs": "COGS",
            "Labour Cost": "Payroll",
            "Occupancy Cost": "Rent",
            "Other Operating Costs": "Other Expenses",
        }
    )
    legacy["Marketing"] = 0
    legacy_buffer = BytesIO()
    legacy.to_excel(legacy_buffer, index=False)
    uploaded_excel = NamedBytesIO(legacy_buffer.getvalue(), "legacy_sample.xlsx")
    read_back = read_uploaded_file(uploaded_excel)
    prepared, errors = validate_and_prepare(read_back)
    if errors:
        raise AssertionError(f"Legacy Excel validation failed: {errors}")

    print("Gate 1 regression passed")
    for result in results:
        print(result)
    print("Legacy schema Excel upload passed")


if __name__ == "__main__":
    run()
