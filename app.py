from __future__ import annotations

from datetime import datetime
from io import BytesIO
from math import isfinite
from textwrap import wrap

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = [
    "Month",
    "Revenue",
    "COGS",
    "Payroll",
    "Rent",
    "Marketing",
    "Other Expenses",
]

MONEY_COLUMNS = [
    "Revenue",
    "COGS",
    "Payroll",
    "Rent",
    "Marketing",
    "Other Expenses",
]


st.set_page_config(
    page_title="Cash Flow Health Check",
    page_icon=":bar_chart:",
    layout="wide",
)


CUSTOM_CSS = """
<style>
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        max-width: 1180px;
    }
    .app-subtitle {
        color: #475569;
        font-size: 1.08rem;
        margin-top: -0.75rem;
        margin-bottom: 1.5rem;
    }
    .metric-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1rem;
        min-height: 112px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }
    .metric-label {
        color: #64748b;
        font-size: 0.85rem;
        font-weight: 600;
        margin-bottom: 0.4rem;
    }
    .metric-value {
        color: #0f172a;
        font-size: 1.55rem;
        font-weight: 750;
        line-height: 1.2;
    }
    .metric-note {
        color: #64748b;
        font-size: 0.8rem;
        margin-top: 0.4rem;
    }
    .score-band {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-left: 6px solid #2563eb;
        border-radius: 8px;
        padding: 1rem 1.1rem;
    }
    .status-healthy { border-left-color: #16a34a; }
    .status-stable { border-left-color: #2563eb; }
    .status-watch { border-left-color: #f59e0b; }
    .status-at-risk { border-left-color: #dc2626; }
    .footer-note {
        color: #64748b;
        font-size: 0.82rem;
        border-top: 1px solid #e2e8f0;
        margin-top: 2rem;
        padding-top: 1rem;
    }
</style>
"""


def money(value: float) -> str:
    return f"${value:,.0f}"


def percent(value: float) -> str:
    return f"{value:.1%}"


def runway_label(value: float, cash_balance: float | None = None) -> str:
    if cash_balance is not None and cash_balance <= 0:
        return "Cash Depleted"
    if value == float("inf"):
        return "Cash Generating"
    if value > 24:
        return "24+ months"
    if value <= 0:
        return "0.0 months"
    return f"{value:.1f} months"


def safe_divide(numerator: float, denominator: float) -> float:
    if denominator == 0 or pd.isna(denominator):
        return 0.0
    result = numerator / denominator
    return float(result) if isfinite(result) else 0.0


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file, engine="openpyxl")


def build_input_template() -> bytes:
    output = BytesIO()
    columns = REQUIRED_COLUMNS
    instructions = [
        ("Month", "Reporting month. Use one row per month, for example Jan 2026 or 2026-01."),
        ("Revenue", "Total business revenue for the month, excluding sales tax, GST, or VAT where applicable."),
        ("COGS", "Direct costs associated with generating revenue."),
        ("Payroll", "Employee wages, salaries, superannuation, and payroll-related costs if applicable."),
        ("Rent", "Premises rent and occupancy costs."),
        ("Marketing", "Advertising and marketing expenditure."),
        ("Other Expenses", "Other operating expenses not included above."),
        ("Currency", "Enter all amounts consistently in the same currency. Do not mix currencies within one file."),
        (
            "Opening Cash Balance",
            "The cash available at the beginning of the first reporting month. Enter this separately in the app. Do not add it as a column in the Monthly Data sheet.",
        ),
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template = pd.DataFrame(columns=columns)
        template["Month"] = pd.date_range("2026-01-01", periods=12, freq="MS").strftime("%b %Y")
        template.to_excel(writer, sheet_name="Monthly Data", index=False, startrow=2)
        pd.DataFrame(instructions, columns=["Field", "Definition"]).to_excel(
            writer, sheet_name="Instructions", index=False
        )

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))

        for sheet in workbook.worksheets:
            header_row = 3 if sheet.title == "Monthly Data" else 1
            sheet.freeze_panes = f"A{header_row + 1}"
            for cell in sheet[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                    max(max_length + 3, 14), 55
                )
        monthly = workbook["Monthly Data"]
        monthly["A1"] = "Replace the example months if needed and enter your own financial data."
        monthly["A1"].font = Font(italic=True, color="475569")
        monthly.merge_cells("A1:G1")
        for row in range(4, 16):
            monthly[f"A{row}"].number_format = "yyyy-mm"
            for col in range(2, 8):
                monthly.cell(row=row, column=col).number_format = "$#,##0"
    return output.getvalue()


def validate_and_prepare(df: pd.DataFrame) -> tuple[pd.DataFrame | None, list[str]]:
    errors: list[str] = []
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        errors.append(f"Missing required columns: {', '.join(missing)}.")
        return None, errors

    df = df[REQUIRED_COLUMNS].copy()
    if len(df) < 3:
        errors.append("At least 3 months of financial data are required.")

    df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
    if df["Month"].isna().any():
        for index in df[df["Month"].isna()].index:
            errors.append(f"Month contains an unreadable value in row {index + 2}.")

    for column in MONEY_COLUMNS:
        original = df[column]
        missing_cells = original.isna() | (original.astype(str).str.strip() == "")
        for index in df[missing_cells].index:
            errors.append(f"{column} is missing a value in row {index + 2}.")
        df[column] = pd.to_numeric(original, errors="coerce")
        non_numeric = df[column].isna() & ~missing_cells
        for index in df[non_numeric].index:
            errors.append(f"{column} contains a non-numeric value in row {index + 2}.")
        negative = df[column] < 0
        if column == "Revenue":
            for index in df[negative].index:
                errors.append(f"Revenue cannot be negative in row {index + 2}.")
        else:
            for index in df[negative].index:
                errors.append(f"{column} cannot be negative in row {index + 2}.")

    valid_months = df["Month"].dropna().dt.to_period("M")
    if valid_months.duplicated().any():
        errors.append("Month appears more than once. Please use one row per reporting month.")

    if errors:
        return None, errors

    df = df.sort_values("Month").reset_index(drop=True)
    return df, []


def calculate_financials(df: pd.DataFrame, opening_cash: float) -> pd.DataFrame:
    result = df.copy()
    result["Gross Profit"] = result["Revenue"] - result["COGS"]
    result["Gross Margin"] = result.apply(
        lambda row: safe_divide(row["Gross Profit"], row["Revenue"]), axis=1
    )
    result["Operating Expenses"] = (
        result["Payroll"] + result["Rent"] + result["Marketing"] + result["Other Expenses"]
    )
    result["Operating Profit"] = result["Gross Profit"] - result["Operating Expenses"]
    result["Operating Margin"] = result.apply(
        lambda row: safe_divide(row["Operating Profit"], row["Revenue"]), axis=1
    )
    result["Estimated Cash Movement"] = result["Operating Profit"]
    result["Estimated Closing Cash Balance"] = opening_cash + result["Estimated Cash Movement"].cumsum()
    result["Break-even Revenue"] = result.apply(
        lambda row: safe_divide(row["Operating Expenses"], row["Gross Margin"])
        if row["Gross Margin"] > 0
        else 0.0,
        axis=1,
    )
    return result


def calculate_burn_and_runway(closing_cash: float, cash_flows: pd.Series) -> tuple[float, float]:
    if (cash_flows > 0).all():
        return 0.0, float("inf")
    burn_months = cash_flows[cash_flows < 0]
    if burn_months.empty:
        return 0.0, float("inf")
    average_burn = abs(float(burn_months.mean()))
    if closing_cash <= 0:
        return average_burn, 0.0
    return average_burn, closing_cash / average_burn


def build_forecast(
    history: pd.DataFrame,
    starting_cash: float,
    revenue_growth: float,
    cogs_percentage: float,
    payroll_growth: float,
    other_opex_growth: float,
) -> pd.DataFrame:
    last = history.iloc[-1]
    rows = []
    prior_revenue = float(last["Revenue"])
    prior_payroll = float(last["Payroll"])
    prior_rent = float(last["Rent"])
    prior_marketing = float(last["Marketing"])
    prior_other = float(last["Other Expenses"])
    prior_month = pd.Timestamp(last["Month"])

    for offset in range(1, 13):
        month = prior_month + pd.DateOffset(months=offset)
        revenue = prior_revenue * ((1 + revenue_growth) ** offset)
        payroll = prior_payroll * ((1 + payroll_growth) ** offset)
        marketing = prior_marketing * ((1 + other_opex_growth) ** offset)
        other_expenses = prior_other * ((1 + other_opex_growth) ** offset)
        rent = prior_rent * ((1 + other_opex_growth) ** offset)
        cogs = revenue * cogs_percentage

        rows.append(
            {
                "Month": month,
                "Revenue": revenue,
                "COGS": cogs,
                "Payroll": payroll,
                "Rent": rent,
                "Marketing": marketing,
                "Other Expenses": other_expenses,
            }
        )

    return calculate_financials(pd.DataFrame(rows), starting_cash)


def build_scenario_from_base(
    base_forecast: pd.DataFrame,
    starting_cash: float,
    scenario_name: str,
    revenue_adjustment: float,
    cogs_percentage: float,
) -> tuple[pd.DataFrame, dict[str, float | str]]:
    scenario = base_forecast[
        ["Month", "Revenue", "Payroll", "Rent", "Marketing", "Other Expenses"]
    ].copy()
    scenario["Revenue"] = scenario["Revenue"] * (1 + revenue_adjustment)
    scenario["COGS"] = scenario["Revenue"] * cogs_percentage
    scenario = scenario[REQUIRED_COLUMNS]
    scenario = calculate_financials(scenario, starting_cash)
    scenario_metrics = summarize_metrics(scenario)
    return scenario, {
        "Scenario": scenario_name,
        "Revenue Adjustment": revenue_adjustment,
        "Revenue": scenario_metrics["revenue"],
        "Operating Profit": scenario_metrics["operating_profit"],
        "Estimated Ending Cash Balance": scenario_metrics["cash_balance"],
        "Cash Runway": runway_label(
            float(scenario_metrics["cash_runway"]), float(scenario_metrics["cash_balance"])
        ),
        "Operating Margin": scenario_metrics["operating_margin"],
    }


def summarize_metrics(financials: pd.DataFrame) -> dict[str, float]:
    latest = financials.iloc[-1]
    total_revenue = float(financials["Revenue"].sum())
    total_operating_profit = float(financials["Operating Profit"].sum())
    latest_revenue = float(latest["Revenue"])
    break_even_revenue = float(latest["Break-even Revenue"])
    average_burn, runway = calculate_burn_and_runway(
        float(latest["Estimated Closing Cash Balance"]), financials["Estimated Cash Movement"]
    )
    return {
        "revenue": total_revenue,
        "gross_margin": float(financials["Gross Profit"].sum() / total_revenue)
        if total_revenue != 0
        else 0.0,
        "operating_profit": total_operating_profit,
        "operating_margin": safe_divide(total_operating_profit, total_revenue),
        "latest_revenue": latest_revenue,
        "cash_balance": float(latest["Estimated Closing Cash Balance"]),
        "cash_runway": runway,
        "average_burn": average_burn,
        "break_even_revenue": break_even_revenue,
        "break_even_buffer": safe_divide(latest_revenue - break_even_revenue, latest_revenue),
    }


def revenue_growth_rate(financials: pd.DataFrame) -> float:
    first = float(financials.iloc[0]["Revenue"])
    last = float(financials.iloc[-1]["Revenue"])
    return safe_divide(last - first, abs(first))


def expense_growth_rate(financials: pd.DataFrame) -> float:
    first = float(financials.iloc[0]["Operating Expenses"])
    last = float(financials.iloc[-1]["Operating Expenses"])
    return safe_divide(last - first, abs(first))


def expense_trend_score(revenue_growth: float, expense_growth: float) -> dict[str, int | str]:
    leverage_gap = revenue_growth - expense_growth
    if (expense_growth < 0 and revenue_growth >= 0) or leverage_gap >= 0.10:
        return score_component(
            15,
            15,
            f"Expense growth is materially below revenue growth; leverage gap is {percent(leverage_gap)}",
        )
    if leverage_gap >= 0.05:
        return score_component(
            12,
            15,
            f"Expense growth is moderately below revenue growth; leverage gap is {percent(leverage_gap)}",
        )
    if leverage_gap >= -0.03:
        return score_component(
            8,
            15,
            f"Expenses and revenue moved at approximately similar rates; leverage gap is {percent(leverage_gap)}",
        )
    if leverage_gap >= -0.10:
        return score_component(
            4,
            15,
            f"Expenses grew moderately faster than revenue; leverage gap is {percent(leverage_gap)}",
        )
    return score_component(
        0,
        15,
        f"Expenses grew materially faster than revenue; leverage gap is {percent(leverage_gap)}",
    )


def score_component(value: int, maximum: int, reason: str) -> dict[str, int | str]:
    return {"Score": value, "Maximum Score": maximum, "Reason": reason}


def score_cash_health(financials: pd.DataFrame, metrics: dict[str, float]) -> tuple[int, str, pd.DataFrame]:
    historical_margin = metrics["operating_margin"]
    rev_growth = revenue_growth_rate(financials)
    exp_growth = expense_growth_rate(financials)
    buffer = metrics["break_even_buffer"]
    runway = metrics["cash_runway"]

    if metrics["cash_balance"] <= 0:
        cash = score_component(0, 30, "Estimated cash resources are depleted")
    elif runway == float("inf"):
        cash = score_component(30, 30, "Business is cash generating")
    elif runway >= 24:
        cash = score_component(30, 30, f"Estimated cash runway is {runway:.1f} months")
    elif runway >= 12:
        cash = score_component(24, 30, f"Estimated cash runway is {runway:.1f} months")
    elif runway >= 6:
        cash = score_component(18, 30, f"Estimated cash runway is {runway:.1f} months")
    elif runway >= 3:
        cash = score_component(10, 30, f"Estimated cash runway is {runway:.1f} months")
    elif runway > 0:
        cash = score_component(5, 30, f"Estimated cash runway is {runway:.1f} months")
    else:
        cash = score_component(0, 30, "Estimated cash balance is depleted")

    if historical_margin >= 0.20:
        margin = score_component(25, 25, f"Operating margin is {percent(historical_margin)}")
    elif historical_margin >= 0.10:
        margin = score_component(20, 25, f"Operating margin is {percent(historical_margin)}")
    elif historical_margin >= 0.05:
        margin = score_component(15, 25, f"Operating margin is {percent(historical_margin)}")
    elif historical_margin >= 0:
        margin = score_component(10, 25, f"Operating margin is {percent(historical_margin)}")
    elif historical_margin >= -0.05:
        margin = score_component(5, 25, f"Operating margin is {percent(historical_margin)}")
    else:
        margin = score_component(0, 25, f"Operating margin is {percent(historical_margin)}")

    if rev_growth >= 0.10:
        revenue = score_component(20, 20, f"Revenue increased by {percent(rev_growth)} over the period")
    elif rev_growth >= 0.03:
        revenue = score_component(16, 20, f"Revenue increased by {percent(rev_growth)} over the period")
    elif rev_growth >= 0:
        revenue = score_component(12, 20, f"Revenue was broadly flat at {percent(rev_growth)} growth")
    elif rev_growth >= -0.05:
        revenue = score_component(6, 20, f"Revenue declined by {percent(abs(rev_growth))} over the period")
    else:
        revenue = score_component(0, 20, f"Revenue declined by {percent(abs(rev_growth))} over the period")

    expense = expense_trend_score(rev_growth, exp_growth)

    if buffer >= 0.25:
        break_even = score_component(10, 10, f"Latest revenue is {percent(buffer)} above break-even")
    elif buffer >= 0.10:
        break_even = score_component(8, 10, f"Latest revenue is {percent(buffer)} above break-even")
    elif buffer >= 0:
        break_even = score_component(5, 10, f"Latest revenue is {percent(buffer)} above break-even")
    elif buffer >= -0.10:
        break_even = score_component(2, 10, f"Latest revenue is {percent(abs(buffer))} below break-even")
    else:
        break_even = score_component(0, 10, f"Latest revenue is {percent(abs(buffer))} below break-even")

    breakdown = pd.DataFrame(
        [
            {"Factor": "Cash Position / Runway", **cash},
            {"Factor": "Operating Margin", **margin},
            {"Factor": "Revenue Trend", **revenue},
            {"Factor": "Expense Trend", **expense},
            {"Factor": "Break-even Buffer", **break_even},
        ]
    )
    score = int(breakdown["Score"].sum())
    if score >= 80:
        label = "Healthy"
    elif score >= 60:
        label = "Stable"
    elif score >= 40:
        label = "Watch"
    else:
        label = "At Risk"
    return score, label, breakdown


def make_cfo_summary(
    financials: pd.DataFrame,
    metrics: dict[str, float],
    score_label: str,
    downside_metrics: dict[str, float] | None = None,
) -> list[str]:
    rev_growth = revenue_growth_rate(financials)
    exp_growth = expense_growth_rate(financials)
    latest = financials.iloc[-1]
    first = financials.iloc[0]
    break_even_delta = metrics["break_even_buffer"]

    if rev_growth > 0.05:
        revenue_line = (
            f"Revenue increased from approximately {money(float(first['Revenue']))} to "
            f"{money(float(latest['Revenue']))} per month ({percent(rev_growth)} total growth)."
        )
    elif rev_growth < -0.05:
        revenue_line = (
            f"Revenue declined from approximately {money(float(first['Revenue']))} to "
            f"{money(float(latest['Revenue']))} per month ({percent(abs(rev_growth))} total decline)."
        )
    else:
        revenue_line = (
            f"Revenue was broadly flat, moving from approximately {money(float(first['Revenue']))} to "
            f"{money(float(latest['Revenue']))} per month."
        )

    if metrics["operating_margin"] >= 0.15:
        margin_note = "strong operating profitability"
    elif metrics["operating_margin"] >= 0.05:
        margin_note = "positive but modest operating profitability"
    elif metrics["operating_margin"] >= 0:
        margin_note = "thin operating profitability"
    else:
        margin_note = "an operating loss"
    profit_line = (
        f"Gross margin is {percent(metrics['gross_margin'])} and operating margin is "
        f"{percent(metrics['operating_margin'])}, indicating {margin_note}."
    )
    cash_line = (
        f"Estimated cash balance is approximately {money(metrics['cash_balance'])}, with "
        f"runway shown as {runway_label(metrics['cash_runway'], metrics['cash_balance'])}."
    )
    if break_even_delta >= 0:
        break_even_line = (
            f"Estimated break-even revenue is approximately {money(metrics['break_even_revenue'])} per month; "
            f"latest monthly revenue is approximately {percent(break_even_delta)} above break-even."
        )
    else:
        break_even_line = (
            f"Estimated break-even revenue is approximately {money(metrics['break_even_revenue'])} per month; "
            f"latest monthly revenue is approximately {percent(abs(break_even_delta))} below break-even."
        )

    downside_line = ""
    if downside_metrics is not None:
        downside_runway = runway_label(
            float(downside_metrics["cash_runway"]), float(downside_metrics["cash_balance"])
        )
        downside_line = (
            f"In the downside scenario, 12-month operating profit is "
            f"{money(float(downside_metrics['operating_profit']))}, estimated ending cash is "
            f"{money(float(downside_metrics['cash_balance']))}, and runway is {downside_runway}."
        )

    if metrics["cash_balance"] <= 0:
        risk = "Primary risk: estimated cash resources are exhausted under the current simplified cash assumptions."
    elif latest["Revenue"] < latest["Break-even Revenue"]:
        risk = "Primary risk: the business is operating below estimated break-even revenue."
    elif metrics["operating_margin"] < 0:
        risk = "Primary risk: operating losses are reducing estimated cash resilience."
    elif exp_growth > rev_growth:
        risk = "Primary risk: operating expenses are growing faster than revenue, weakening operating leverage."
    else:
        risk = "Primary risk: preserving margin quality and downside resilience as the business grows."

    if score_label == "Healthy":
        action = "Recommended action: preserve operating leverage, maintain margin discipline, and deploy surplus cash carefully."
    elif score_label == "Stable":
        action = "Recommended action: monitor margin, cost growth, and downside resilience before committing to additional fixed costs."
    elif score_label == "Watch":
        action = "Recommended action: review pricing, reduce discretionary costs, improve gross margin, tighten working capital discipline, and preserve cash."
    else:
        action = "Recommended action: immediate management attention is required to reduce cash outflows, restore gross margin, and protect liquidity."

    return [
        line
        for line in [revenue_line, profit_line, cash_line, break_even_line, downside_line, risk, action]
        if line
    ]


def format_excel_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_length + 3, 14), 42
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = sheet.cell(row=1, column=cell.column).value or ""
                    if "Margin" in header or "Growth" in header or "%" in header or "Buffer" in header or "Adjustment" in header:
                        cell.number_format = "0.0%"
                    elif "Score" not in header:
                        cell.number_format = "$#,##0"
                elif sheet.cell(row=1, column=cell.column).value == "Month":
                    cell.number_format = "mmm yyyy"


def export_excel(
    history: pd.DataFrame,
    forecast: pd.DataFrame,
    scenarios: pd.DataFrame,
    score_breakdown: pd.DataFrame,
    assumptions: dict[str, float],
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        history.to_excel(writer, sheet_name="Historical Analysis", index=False)
        forecast.to_excel(writer, sheet_name="12-Month Forecast", index=False)
        scenarios.to_excel(writer, sheet_name="Scenario Analysis", index=False)
        score_breakdown.to_excel(writer, sheet_name="Health Score", index=False)
        pd.DataFrame(
            [{"Assumption": key, "Value": value} for key, value in assumptions.items()]
        ).to_excel(writer, sheet_name="Assumptions", index=False)
        format_excel_workbook(writer)
    return output.getvalue()


def pdf_text_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(
    content: list[str],
    x: float,
    y: float,
    text: str,
    size: int = 10,
    color: tuple[float, float, float] = (0.05, 0.07, 0.10),
) -> None:
    content.append(f"{color[0]:.3f} {color[1]:.3f} {color[2]:.3f} rg")
    content.append("BT")
    content.append(f"/F1 {size} Tf")
    content.append(f"1 0 0 1 {x:.1f} {y:.1f} Tm")
    content.append(f"({pdf_text_escape(text)}) Tj")
    content.append("ET")


def pdf_wrapped_text(
    content: list[str],
    x: float,
    y: float,
    text: str,
    size: int = 10,
    width: int = 86,
    line_height: float = 14,
    color: tuple[float, float, float] = (0.05, 0.07, 0.10),
) -> float:
    for part in wrap(text, width=width) or [""]:
        pdf_text(content, x, y, part, size, color=color)
        y -= line_height
    return y


def pdf_rect(
    content: list[str],
    x: float,
    y: float,
    width: float,
    height: float,
    fill: tuple[float, float, float] | None = None,
    stroke: tuple[float, float, float] | None = (0.82, 0.86, 0.91),
) -> None:
    content.append("q")
    if fill:
        content.append(f"{fill[0]:.3f} {fill[1]:.3f} {fill[2]:.3f} rg")
        content.append(f"{x:.1f} {y:.1f} {width:.1f} {height:.1f} re f")
    if stroke:
        content.append(f"{stroke[0]:.3f} {stroke[1]:.3f} {stroke[2]:.3f} RG")
        content.append(f"{x:.1f} {y:.1f} {width:.1f} {height:.1f} re S")
    content.append("Q")


def pdf_footer(content: list[str], page_number: int) -> None:
    content.append("0.82 0.86 0.91 RG")
    content.append("54 44 504 0 m S")
    pdf_text(
        content,
        72,
        26,
        f"Cash Flow Health Check | Page {page_number} of 3",
        9,
        color=(0.28, 0.33, 0.40),
    )


def pdf_page_stream(content: list[str], page_number: int) -> bytes:
    pdf_footer(content, page_number)
    return "\n".join(content).encode("latin-1", errors="replace")


def build_report_pdf(
    metrics: dict[str, float],
    score: int,
    score_label: str,
    forecast_metrics: dict[str, float],
    scenarios: pd.DataFrame,
    summary_lines: list[str],
) -> bytes:
    generated_date = datetime.now().strftime("%Y-%m-%d")
    pages: list[list[str]] = []

    page1: list[str] = []
    pdf_text(page1, 54, 742, "Cash Flow Health Report", 20)
    pdf_text(page1, 54, 718, "Financial Health and 12-Month Scenario Analysis", 12, color=(0.28, 0.33, 0.40))
    pdf_text(page1, 54, 698, f"Report generated: {generated_date}", 9, color=(0.35, 0.39, 0.46))
    pdf_rect(page1, 54, 592, 504, 78, fill=(0.94, 0.97, 1.0))
    pdf_text(page1, 76, 640, "Cash Flow Health Score", 11, color=(0.23, 0.28, 0.34))
    pdf_text(page1, 76, 610, f"{score}/100", 24)
    pdf_text(page1, 330, 640, "Status", 11, color=(0.23, 0.28, 0.34))
    pdf_text(page1, 330, 610, score_label, 24)
    snapshot = [
        ("Estimated Cash Balance", money(metrics["cash_balance"])),
        ("Cash Runway", runway_label(metrics["cash_runway"], metrics["cash_balance"])),
        ("Revenue", money(metrics["revenue"])),
        ("Gross Margin", percent(metrics["gross_margin"])),
        ("Operating Margin", percent(metrics["operating_margin"])),
        ("Break-even Revenue", money(metrics["break_even_revenue"])),
        ("Break-even Buffer", percent(metrics["break_even_buffer"])),
    ]
    x_positions = [54, 312]
    y = 526
    for idx, (label, value) in enumerate(snapshot):
        x = x_positions[idx % 2]
        if idx % 2 == 0 and idx > 0:
            y -= 78
        pdf_rect(page1, x, y, 246, 56, fill=(1, 1, 1))
        pdf_text(page1, x + 14, y + 34, label, 9, color=(0.28, 0.33, 0.40))
        pdf_text(page1, x + 14, y + 13, value, 14)
    y_note = 124
    pdf_text(page1, 54, y_note + 34, "Methodology note", 10, color=(0.23, 0.28, 0.34))
    pdf_wrapped_text(
        page1,
        54,
        y_note + 16,
        "Estimated cash movement uses operating profit as a simplified proxy for operating cash flow. It does not include working capital movements, capital expenditure, financing, tax payments, or owner distributions.",
        8,
        width=112,
        line_height=11,
        color=(0.28, 0.33, 0.40),
    )
    pages.append(page1)

    page2: list[str] = []
    pdf_text(page2, 54, 742, "12-Month Forecast and Scenario Analysis", 17)
    pdf_rect(page2, 54, 666, 504, 34, fill=(0.12, 0.23, 0.47), stroke=None)
    headers = ["Scenario", "Revenue", "Operating Profit", "Estimated Ending Cash", "Cash Runway"]
    cols = [66, 162, 258, 370, 474]
    for x, header in zip(cols, headers):
        pdf_text(page2, x, 678, header, 8, color=(1.0, 1.0, 1.0))
    row_y = 642
    for _, row in scenarios.iterrows():
        pdf_rect(page2, 54, row_y - 8, 504, 22, fill=(1, 1, 1), stroke=(0.90, 0.92, 0.95))
        pdf_text(page2, cols[0], row_y, str(row["Scenario"]), 9)
        pdf_text(page2, cols[1], row_y, money(float(row["Revenue"])), 9)
        pdf_text(page2, cols[2], row_y, money(float(row["Operating Profit"])), 9)
        pdf_text(page2, cols[3], row_y, money(float(row["Estimated Ending Cash Balance"])), 9)
        pdf_text(page2, cols[4], row_y, str(row["Cash Runway"]), 9)
        row_y -= 28
    pdf_text(page2, 54, 520, "Estimated Ending Cash by Scenario", 13)
    chart_x, chart_y, chart_w, bar_h = 190, 434, 300, 22
    max_cash = max(abs(float(value)) for value in scenarios["Estimated Ending Cash Balance"]) or 1
    for idx, (_, row) in enumerate(scenarios.iterrows()):
        y_bar = chart_y - idx * 48
        cash_value = float(row["Estimated Ending Cash Balance"])
        width = abs(cash_value) / max_cash * chart_w
        fill = (0.15, 0.39, 0.92) if cash_value >= 0 else (0.86, 0.15, 0.15)
        pdf_text(page2, 66, y_bar + 6, str(row["Scenario"]), 9)
        pdf_rect(page2, chart_x, y_bar, chart_w, bar_h, fill=(0.96, 0.97, 0.98), stroke=None)
        pdf_rect(page2, chart_x, y_bar, width, bar_h, fill=fill, stroke=None)
        pdf_text(page2, chart_x + chart_w + 12, y_bar + 6, money(cash_value), 9)
    pdf_text(page2, 54, 250, "12-Month Estimated Cash Balance Trend", 13)
    pdf_wrapped_text(
        page2,
        54,
        226,
        f"The base forecast ends with estimated cash of {money(forecast_metrics['cash_balance'])}. Detailed monthly forecast values are available in the Excel export and in the application dashboard.",
        10,
        width=100,
        color=(0.23, 0.28, 0.34),
    )
    pages.append(page2)

    risk_lines = [line for line in summary_lines if line.startswith("Primary risk:")]
    action_lines = [line for line in summary_lines if line.startswith("Recommended action:")]
    key_lines = [line for line in summary_lines if line not in risk_lines + action_lines]
    page3: list[str] = []
    pdf_text(page3, 54, 742, "CFO Summary", 17)
    pdf_text(page3, 54, 706, "Key Findings", 13)
    y = 682
    for line in key_lines[:5]:
        y = pdf_wrapped_text(page3, 70, y, f"- {line}", 10, width=92)
        y -= 6
    pdf_text(page3, 54, 390, "Primary Financial Risk", 13)
    pdf_rect(page3, 54, 304, 504, 68, fill=(1.0, 0.95, 0.95), stroke=(0.95, 0.65, 0.65))
    risk_text = risk_lines[0].replace("Primary risk: ", "") if risk_lines else "No primary risk identified."
    pdf_wrapped_text(page3, 76, 346, risk_text, 10, width=92)
    pdf_text(page3, 54, 260, "Recommended Management Actions", 13)
    action_text = (
        action_lines[0].replace("Recommended action: ", "")
        if action_lines
        else "Review the dashboard indicators and scenario results before making management decisions."
    )
    pdf_wrapped_text(page3, 70, 236, f"1. {action_text}", 10, width=92)
    pdf_text(page3, 54, 118, "Limitations and Disclaimer", 11, color=(0.23, 0.28, 0.34))
    pdf_wrapped_text(
        page3,
        54,
        100,
        "This tool provides an indicative financial analysis based on information supplied by the user. It is not accounting, tax, audit, investment, or financial advice. Results should be reviewed alongside actual cash-flow records and professional advice where appropriate.",
        8,
        width=112,
        line_height=10,
        color=(0.28, 0.33, 0.40),
    )
    pages.append(page3)

    streams = [pdf_page_stream(page, index) for index, page in enumerate(pages, start=1)]
    page_objects = []
    content_object_start = 3 + len(streams) + 1
    for idx, _ in enumerate(streams, start=1):
        content_ref = content_object_start + idx - 1
        page_objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 3 0 R >> >> /Contents {content_ref} 0 R >>".encode()
        )

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{' '.join(f'{i} 0 R' for i in range(4, 4 + len(streams)))}] /Count {len(streams)} >>".encode(),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        *page_objects,
        *[
            b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
            for stream in streams
        ],
    ]
    pdf = BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = []
    for index, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f"{index} 0 obj\n".encode())
        pdf.write(obj)
        pdf.write(b"\nendobj\n")
    xref = pdf.tell()
    pdf.write(f"xref\n0 {len(objects) + 1}\n".encode())
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets:
        pdf.write(f"{offset:010d} 00000 n \n".encode())
    pdf.write(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode()
    )
    return pdf.getvalue()


def metric_card(label: str, value: str, note: str = "") -> None:
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def style_chart(fig):
    fig.update_layout(
        template="plotly_white",
        margin=dict(l=10, r=10, t=36, b=10),
        hovermode="x unified",
        height=330,
    )
    fig.update_xaxes(title=None)
    return fig


st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
st.title("Cash Flow Health Check")
st.markdown(
    '<div class="app-subtitle">Understand your cash position, test scenarios, and see your next 12 months.</div>',
    unsafe_allow_html=True,
)
st.caption(
    "Built for small business owners who want a clearer view of profitability, cash resilience, and downside risk."
)

with st.sidebar:
    st.header("Inputs")
    st.markdown("1. Download the template")
    st.markdown("2. Enter your monthly financial data")
    st.markdown("3. Upload the completed CSV or Excel file")
    st.download_button(
        "Download Input Template",
        data=build_input_template(),
        file_name="cash_flow_health_check_input_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    uploaded = st.file_uploader("Upload CSV or Excel", type=["csv", "xlsx", "xls"])
    opening_cash = st.number_input(
        "Opening Cash Balance",
        value=0.0,
        step=1000.0,
        format="%.2f",
        help="Enter the cash available at the beginning of your first reporting month.",
    )
    st.caption(
        "Required columns: Month, Revenue, COGS (direct costs), Payroll, Rent, Marketing, Other Expenses."
    )

if opening_cash < 0:
    st.error("Opening Cash Balance cannot be negative.")
    st.stop()

if uploaded is None:
    st.info("Upload your completed financial data file to generate your Cash Flow Health Check.")
    st.stop()

try:
    raw_df = read_uploaded_file(uploaded)
except Exception as exc:
    st.error(f"Could not read the uploaded file: {exc}")
    st.stop()

prepared_df, validation_errors = validate_and_prepare(raw_df)
if validation_errors:
    st.error("Please fix the uploaded file before continuing.")
    for error in validation_errors:
        st.write(f"- {error}")
    st.stop()

history = calculate_financials(prepared_df, opening_cash)
metrics = summarize_metrics(history)
score, score_label, score_breakdown = score_cash_health(history, metrics)

st.subheader("Dashboard")
st.caption(
    "Estimated cash movement uses operating profit as a simplified proxy for operating cash flow. "
    "It does not include working capital movements, capital expenditure, financing, tax payments, or owner distributions."
)
status_class = score_label.lower().replace(" ", "-")
status_interpretations = {
    "Healthy": "Your business is currently showing strong operating and estimated cash-flow fundamentals.",
    "Stable": "Your business is broadly stable, but some indicators should be monitored.",
    "Watch": "Your business shows financial pressure that should be addressed before liquidity weakens further.",
    "At Risk": "Your business is showing material financial stress and requires immediate management attention.",
}
st.markdown(
    f"""
    <div class="score-band status-{status_class}">
        <div class="metric-label">Cash Flow Health Score</div>
        <div class="metric-value">{score}/100 - {score_label}</div>
        <div class="metric-note">{status_interpretations[score_label]}</div>
    </div>
    """,
    unsafe_allow_html=True,
)
if metrics["cash_balance"] <= 0:
    st.warning(
        "Estimated cash resources are exhausted within the analysed period under the current simplified cash assumptions."
    )

kpi_cols = st.columns(3)
with kpi_cols[0]:
    metric_card("Estimated Cash Position", money(metrics["cash_balance"]), "Latest estimated closing balance")
with kpi_cols[1]:
    metric_card(
        "Cash Runway",
        runway_label(metrics["cash_runway"], metrics["cash_balance"]),
        "Based on estimated cash movement",
    )
with kpi_cols[2]:
    metric_card("Operating Margin", percent(metrics["operating_margin"]), "Historical weighted average")

kpi_cols = st.columns(3)
with kpi_cols[0]:
    metric_card("Break-even Buffer", percent(metrics["break_even_buffer"]), "Latest revenue above break-even")
with kpi_cols[1]:
    metric_card("Revenue", money(metrics["revenue"]), "Historical total")
with kpi_cols[2]:
    metric_card("Gross Margin", percent(metrics["gross_margin"]), "Historical weighted average")

with st.expander("View score breakdown"):
    st.dataframe(score_breakdown, use_container_width=True, hide_index=True)
    st.caption(f"Component scores total {int(score_breakdown['Score'].sum())}/100.")
    st.caption(
        "Expense Trend scoring compares revenue growth with operating expense growth: "
        "15 points for a leverage gap of at least 10 percentage points or declining expenses with stable/growing revenue; "
        "12 for at least 5 points; 8 for approximately similar growth within -3 points; "
        "4 when expenses grow up to 10 points faster; 0 when expenses grow materially faster."
    )

diagnostic_cols = st.columns(3)
with diagnostic_cols[0]:
    metric_card("Break-even Revenue", money(metrics["break_even_revenue"]), "Latest monthly estimate")
with diagnostic_cols[1]:
    if metrics["average_burn"] > 0:
        metric_card("Average Monthly Cash Burn", money(metrics["average_burn"]), "Average of negative estimated cash-movement months")
    else:
        metric_card("Cash Burn", "$0 / No burn", "Estimated cash movement is not negative")
with diagnostic_cols[2]:
    metric_card("Break-even Buffer", percent(metrics["break_even_buffer"]), "Latest revenue above break-even")

chart_cols = st.columns(3)
with chart_cols[0]:
    fig = px.bar(history, x="Month", y="Revenue", title="Monthly Revenue")
    st.plotly_chart(style_chart(fig), use_container_width=True)
with chart_cols[1]:
    fig = px.bar(history, x="Month", y="Operating Profit", title="Monthly Operating Profit")
    st.plotly_chart(style_chart(fig), use_container_width=True)
with chart_cols[2]:
    fig = px.line(
        history,
        x="Month",
        y="Estimated Closing Cash Balance",
        markers=True,
        title="Monthly Estimated Cash Balance",
    )
    st.plotly_chart(style_chart(fig), use_container_width=True)

st.subheader("12-month Forecast")
assumption_cols = st.columns(4)
with assumption_cols[0]:
    revenue_growth = st.number_input("Monthly revenue growth", value=3.0, step=0.5, format="%.2f") / 100
with assumption_cols[1]:
    cogs_percentage = st.number_input("COGS percentage", value=40.0, min_value=0.0, max_value=100.0, step=1.0) / 100
with assumption_cols[2]:
    payroll_growth = st.number_input("Payroll growth", value=1.0, step=0.5, format="%.2f") / 100
with assumption_cols[3]:
    other_opex_growth = st.number_input("Other operating expense growth", value=1.0, step=0.5, format="%.2f") / 100

forecast = build_forecast(
    history,
    metrics["cash_balance"],
    revenue_growth,
    cogs_percentage,
    payroll_growth,
    other_opex_growth,
)
forecast_metrics = summarize_metrics(forecast)

forecast_cols = st.columns(3)
with forecast_cols[0]:
    metric_card("Forecast Revenue", money(forecast_metrics["revenue"]), "Next 12 months")
with forecast_cols[1]:
    metric_card("Forecast Operating Profit", money(forecast_metrics["operating_profit"]), "Next 12 months")
with forecast_cols[2]:
    metric_card("Forecast Estimated Cash Balance", money(forecast_metrics["cash_balance"]), "Month 12")

forecast_runway_cols = st.columns(2)
with forecast_runway_cols[0]:
    metric_card(
        "Forecast Cash Runway",
        runway_label(forecast_metrics["cash_runway"], forecast_metrics["cash_balance"]),
        "Based on forecast estimated cash movement",
    )
with forecast_runway_cols[1]:
    if forecast_metrics["average_burn"] > 0:
        metric_card("Forecast Average Monthly Cash Burn", money(forecast_metrics["average_burn"]), "Average of negative forecast months")
    else:
        metric_card("Forecast Cash Burn", "$0 / No burn", "Forecast is cash generating")

fig = px.line(
    forecast,
    x="Month",
    y=["Revenue", "Operating Profit", "Estimated Closing Cash Balance"],
    title="12-month Forecast",
)
st.plotly_chart(style_chart(fig), use_container_width=True)

with st.expander("View forecast table"):
    st.dataframe(forecast, use_container_width=True)

st.subheader("Scenario Analysis")
st.caption("Scenario adjustments are one-time revenue adjustments versus the 12-month Base Case forecast, not compounded monthly growth rates.")
scenario_cols = st.columns(2)
with scenario_cols[0]:
    downside_adjustment = st.number_input("Downside Case revenue adjustment", value=-15.0, step=1.0, format="%.1f") / 100
with scenario_cols[1]:
    upside_adjustment = st.number_input("Upside Case revenue adjustment", value=15.0, step=1.0, format="%.1f") / 100

scenario_rows = []
scenario_details = {}
for scenario_name, adjustment in [
    ("Base Case", 0.0),
    ("Downside Case", downside_adjustment),
    ("Upside Case", upside_adjustment),
]:
    scenario_forecast, scenario_row = build_scenario_from_base(
        forecast,
        metrics["cash_balance"],
        scenario_name,
        adjustment,
        cogs_percentage,
    )
    scenario_details[scenario_name] = {
        "forecast": scenario_forecast,
        "metrics": summarize_metrics(scenario_forecast),
    }
    scenario_rows.append(scenario_row)

scenarios = pd.DataFrame(scenario_rows)
st.dataframe(
    scenarios.style.format(
        {
            "Revenue Adjustment": "{:.1%}",
            "Revenue": "${:,.0f}",
            "Operating Profit": "${:,.0f}",
            "Estimated Ending Cash Balance": "${:,.0f}",
            "Operating Margin": "{:.1%}",
        }
    ),
    use_container_width=True,
)

fig = px.bar(
    scenarios,
    x="Scenario",
    y=["Revenue", "Operating Profit", "Estimated Ending Cash Balance"],
    barmode="group",
    title="Scenario Outcomes",
)
st.plotly_chart(style_chart(fig), use_container_width=True)

downside_metrics = scenario_details["Downside Case"]["metrics"]
if downside_metrics["cash_balance"] <= 0:
    st.info("The selected downside scenario indicates a potential funding shortfall.")
elif downside_metrics["operating_profit"] < 0:
    st.info("The selected downside scenario materially weakens profitability and estimated cash resilience.")
else:
    st.info("The business remains estimated cash-generative under the selected downside scenario.")

st.subheader("CFO Summary")
summary_lines = make_cfo_summary(
    history,
    metrics,
    score_label,
    downside_metrics,
)
for line in summary_lines:
    st.write(f"- {line}")

st.subheader("Exports")
export_cols = st.columns(2)
with export_cols[0]:
    assumptions = {
        "Opening Cash Balance": opening_cash,
        "Forecast Revenue Growth": revenue_growth,
        "COGS %": cogs_percentage,
        "Payroll Growth": payroll_growth,
        "Other Operating Expense Growth": other_opex_growth,
        "Downside Adjustment": downside_adjustment,
        "Upside Adjustment": upside_adjustment,
    }
    st.download_button(
        "Download Forecast Excel",
        data=export_excel(history, forecast, scenarios, score_breakdown, assumptions),
        file_name=f"cash_flow_forecast_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with export_cols[1]:
    st.download_button(
        "Download Summary PDF",
        data=build_report_pdf(metrics, score, score_label, forecast_metrics, scenarios, summary_lines),
        file_name=f"cash_flow_summary_{datetime.now().strftime('%Y%m%d')}.pdf",
        mime="application/pdf",
    )

st.markdown(
    '<div class="footer-note">This tool provides an indicative financial analysis based on information supplied by the user. It is not accounting, tax, audit, investment, or financial advice. Results should be reviewed alongside actual cash-flow records and professional advice where appropriate.</div>',
    unsafe_allow_html=True,
)
